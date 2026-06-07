import asyncio
import os
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import json
import datetime
from openai import OpenAI
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    MessageEntity,
    User as TGUser,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    ChatJoinRequestHandler,
    ChatMemberHandler,
    filters,
)

TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
ADMIN_ID = int(os.environ["ADMIN_TELEGRAM_ID"])
ADMIN_IDS = {ADMIN_ID, 6345192246}  # @LosevM (основной) + @losev70
CHAT_ID = int(os.environ["CHAT_TELEGRAM_ID"])
STATS_THREAD_ID = int(os.environ["STATS_THREAD_ID"]) if os.environ.get("STATS_THREAD_ID") else None
DIGEST_THREAD_ID = int(os.environ["DIGEST_THREAD_ID"]) if os.environ.get("DIGEST_THREAD_ID") else None
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
INTRO_THREAD_ID = int(os.environ["INTRO_THREAD_ID"]) if os.environ.get("INTRO_THREAD_ID") else None
GENERAL_THREAD_ID = int(os.environ["GENERAL_THREAD_ID"]) if os.environ.get("GENERAL_THREAD_ID") else None
VERIFY_THREAD_ID = int(os.environ["VERIFY_THREAD_ID"]) if os.environ.get("VERIFY_THREAD_ID") else None

# Числовой ID чата для ссылок вида t.me/c/{id}/{msg_id}
CHAT_LINK_ID = str(CHAT_ID).replace("-100", "").lstrip("-")

HOUSE_PHOTO = os.path.join(os.path.dirname(__file__), "house.jpg")

users = {}
waiting_photo = set()
waiting_nudge_photo = set()
pending_forward = {}  # admin_id -> {"user_id": int, "name": str}
DB_FILE = os.path.join(os.path.dirname(__file__), "residents.json")
MESSAGES_FILE = os.path.join(os.path.dirname(__file__), "daily_messages.json")
UNREGISTERED_FILE = os.path.join(os.path.dirname(__file__), "unregistered.json")
DIGEST_ARCHIVE_FILE = os.path.join(os.path.dirname(__file__), "digest_archive.json")
JOIN_REQUESTS_FILE = os.path.join(os.path.dirname(__file__), "join_requests.json")
JOIN_REQUESTS_META_FILE = os.path.join(os.path.dirname(__file__), "join_requests_meta.json")
LAST_STATS_MSG_FILE = os.path.join(os.path.dirname(__file__), "last_stats_msg.json")
LAST_VERIFY_MSG_FILE = os.path.join(os.path.dirname(__file__), "last_verify_msg.json")


def load_last_stats_msg() -> int | None:
    if os.path.exists(LAST_STATS_MSG_FILE):
        with open(LAST_STATS_MSG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("message_id")
    return None


def save_last_stats_msg(message_id: int):
    with open(LAST_STATS_MSG_FILE, "w", encoding="utf-8") as f:
        json.dump({"message_id": message_id}, f)


def load_last_verify_msg() -> int | None:
    if os.path.exists(LAST_VERIFY_MSG_FILE):
        with open(LAST_VERIFY_MSG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("message_id")
    return None


def save_last_verify_msg(message_id: int):
    with open(LAST_VERIFY_MSG_FILE, "w", encoding="utf-8") as f:
        json.dump({"message_id": message_id}, f)


def load_join_requests() -> dict:
    """Загружает заявки на вступление из файла (ключи — строки, конвертируем в int)."""
    if os.path.exists(JOIN_REQUESTS_FILE):
        with open(JOIN_REQUESTS_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        return {int(k): v for k, v in raw.items()}
    return {}


def save_join_requests(data: dict):
    with open(JOIN_REQUESTS_FILE, "w", encoding="utf-8") as f:
        json.dump({str(k): v for k, v in data.items()}, f, ensure_ascii=False, indent=2)


join_requests: dict = load_join_requests()  # user_id -> chat_id, хранит заявки на вступление


def load_daily_messages():
    """Загружает сообщения за сегодня из файла."""
    today = datetime.date.today().isoformat()
    if os.path.exists(MESSAGES_FILE):
        with open(MESSAGES_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if data.get("date") == today:
            return data.get("messages", [])
    return []


def save_daily_messages(messages):
    """Сохраняет сообщения за сегодня в файл."""
    today = datetime.date.today().isoformat()
    with open(MESSAGES_FILE, "w", encoding="utf-8") as f:
        json.dump({"date": today, "messages": messages}, f, ensure_ascii=False, indent=2)


def clear_daily_messages():
    """Очищает файл после отправки дайджеста."""
    today = datetime.date.today().isoformat()
    with open(MESSAGES_FILE, "w", encoding="utf-8") as f:
        json.dump({"date": today, "messages": []}, f, ensure_ascii=False)


# Загружаем сообщения за сегодня при старте (переживают перезапуск)
daily_messages = load_daily_messages()

def load_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

residents = load_db()

floor_flats = {
    2: range(312, 322),
    3: range(322, 332),
    4: range(332, 342),
    5: range(342, 352),
    6: range(352, 362),
    7: range(362, 372),
    8: range(372, 382),
    9: range(382, 392),
    10: range(392, 402),
    11: range(402, 412),
    12: range(412, 422),
    13: range(422, 432),
    14: range(432, 442),
    15: range(442, 452),
    16: range(452, 462),
    17: range(462, 472),
    18: range(472, 482),
    19: range(482, 492),
    20: range(492, 502),
    21: range(502, 512),
    22: range(512, 522),
    23: range(522, 532),
}


def save_db(data):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_unregistered():
    if os.path.exists(UNREGISTERED_FILE):
        with open(UNREGISTERED_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_unregistered(data):
    with open(UNREGISTERED_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def is_in_db(user_id: int) -> bool:
    """Проверяет, есть ли пользователь в базе жильцов."""
    db = load_db()
    if str(user_id) in db:
        return True
    return any(v.get("user_id") == user_id for v in db.values())


def get_universal_tags(db: dict, exclude_key: str | None = None) -> list[str]:
    """Возвращает теги которые стоят у ВСЕХ жильцов в базе.
    Если хоть у одного нет тегов вообще — возвращает [].
    exclude_key — ключ записи которую не учитывать (новый жилец ещё не добавлен)."""
    records = [v for k, v in db.items() if k != exclude_key]
    if not records:
        return []
    sets = [set(v.get("tags", [])) for v in records]
    if any(len(s) == 0 for s in sets):
        return []
    common = sets[0].intersection(*sets[1:])
    return sorted(common)


def load_digest_archive() -> dict:
    if os.path.exists(DIGEST_ARCHIVE_FILE):
        with open(DIGEST_ARCHIVE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_digest_to_archive(date_str: str, text: str):
    """Сохраняет текст дайджеста в архив по ключу даты (YYYY-MM-DD)."""
    archive = load_digest_archive()
    archive[date_str] = text
    with open(DIGEST_ARCHIVE_FILE, "w", encoding="utf-8") as f:
        json.dump(archive, f, ensure_ascii=False, indent=2)


WELCOME_TEXT = (
    "Привет 👋\n"
    "Для вступления в чат сначала пройдите небольшую проверку.\n\n"
    "Ты из нашего дома 6.2.2, красная высотка?\n\n"
    "<i>Если кнопки не работают — напишите <a href=\"https://t.me/losev70\">@losev70</a></i>"
)

WELCOME_KEYBOARD = InlineKeyboardMarkup([
    [
        InlineKeyboardButton("🟢 Да", callback_data="yes_house"),
        InlineKeyboardButton("🔴 Нет", callback_data="no_house"),
    ]
])


async def send_welcome(chat_id: int, context: ContextTypes.DEFAULT_TYPE):
    if os.path.exists(HOUSE_PHOTO):
        with open(HOUSE_PHOTO, "rb") as photo:
            await context.bot.send_photo(
                chat_id=chat_id,
                photo=photo,
                caption=WELCOME_TEXT,
                parse_mode="HTML",
                reply_markup=WELCOME_KEYBOARD
            )
    else:
        await context.bot.send_message(
            chat_id=chat_id,
            text=WELCOME_TEXT,
            parse_mode="HTML",
            reply_markup=WELCOME_KEYBOARD
        )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    # Если пришли по кнопке из /postverify (?start=verify) — сразу запускаем верификацию.
    # Ссылка размещена только внутри группы, поэтому кликнуть может только участник чата.
    # Не проверяем get_chat_member — у части пользователей это падает из-за настроек приватности.
    uname = f"@{user.username}" if user.username else "нет username"
    user_link = f"https://t.me/{user.username}" if user.username else f"tg://user?id={user.id}"

    if context.args and context.args[0] == "verify":
        # Проверяем — уже в базе?
        _db = load_db()
        _existing_key = None
        if str(user.id) in _db:
            _existing_key = str(user.id)
        elif any(v.get("user_id") == user.id for v in _db.values()):
            _existing_key = next(k for k, v in _db.items() if v.get("user_id") == user.id)
        elif user.username:
            for k, v in _db.items():
                if v.get("telegram", "").lower() == f"@{user.username}".lower():
                    _existing_key = k
                    break
        if _existing_key:
            _flat = _db[_existing_key].get("flat", 0)
            if _flat:
                _is_join = user.id in join_requests
                _msg = (
                    f"✅ {user.full_name}, ты уже зарегистрирован как житель кв.\u00a0{_flat}.\n"
                    + ("Заявка передана администратору — скоро откроют доступ." if _is_join
                       else "Если что-то изменилось — обратись к администратору.")
                )
                await context.bot.send_message(chat_id=user.id, text=_msg)
                # Если заявка на вступление — уведомляем администратора
                if _is_join:
                    _floor = _db[_existing_key].get("floor", "?")
                    _tg = _db[_existing_key].get("telegram", "нет")
                    _tg_link = _db[_existing_key].get("telegram_link") or f"tg://user?id={user.id}"
                    _keyboard = InlineKeyboardMarkup([
                        [
                            InlineKeyboardButton("✅ Принять", callback_data=f"approve_{user.id}"),
                            InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_{user.id}"),
                        ]
                    ])
                    try:
                        await context.bot.send_message(
                            chat_id=ADMIN_ID,
                            text=(
                                f"📋 <b>Заявка на вступление (уже в базе)</b>\n\n"
                                f"👤 <a href=\"{_tg_link}\">{user.full_name}</a>\n"
                                f"🏢 Этаж: {_floor}\n"
                                f"🚪 Квартира: {_flat}\n"
                                f"🔗 {_tg}"
                            ),
                            parse_mode="HTML",
                            reply_markup=_keyboard,
                        )
                    except Exception:
                        pass
                return

        users[user.id] = {"step": 1}
        await context.bot.send_message(
            chat_id=user.id,
            text="Напиши свой этаж (от 2 до 23)."
        )
        try:
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=(
                    f"🔑 <b>Жилец начал верификацию</b>\n"
                    f"👤 <a href=\"{user_link}\">{user.full_name}</a>\n"
                    f"Username: {uname}\n"
                    f"ID: <code>{user.id}</code>"
                ),
                parse_mode="HTML",
            )
        except Exception as e:
            print(f"[start/verify] Ошибка уведомления админа: {e}")
        return

    # Обычный /start — стандартное приветствие
    await send_welcome(user.id, context)
    # Уведомляем администратора
    try:
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=(
                f"👤 <b>Новый пользователь открыл бота</b>\n"
                f"👤 <a href=\"{user_link}\">{user.full_name}</a>\n"
                f"Username: {uname}\n"
                f"ID: <code>{user.id}</code>"
            ),
            parse_mode="HTML",
        )
    except Exception as e:
        print(f"[start] Ошибка уведомления админа: {e}")


async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    data = query.data
    print(f"[button] user_id={user_id} name={query.from_user.full_name!r} data={data!r}")

    if data == "yes_house":
        # Проверяем — уже в базе с квартирой?
        _db = load_db()
        _ekey = None
        _u = query.from_user
        if str(user_id) in _db:
            _ekey = str(user_id)
        elif any(v.get("user_id") == user_id for v in _db.values()):
            _ekey = next(k for k, v in _db.items() if v.get("user_id") == user_id)
        elif _u.username:
            for k, v in _db.items():
                if v.get("telegram", "").lower() == f"@{_u.username}".lower():
                    _ekey = k
                    break
        if _ekey and _db[_ekey].get("flat", 0):
            _flat = _db[_ekey].get("flat")
            _floor = _db[_ekey].get("floor", "?")
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=(
                        f"✅ {_u.full_name}, ты уже зарегистрирован как житель кв.\u00a0{_flat}.\n"
                        f"Заявка передана администратору — скоро откроют доступ."
                    ),
                )
            except Exception:
                pass
            # Если это заявка на вступление — уведомляем администратора с кнопками одобрения
            if user_id in join_requests:
                _tg = _db[_ekey].get("telegram", "нет")
                _tg_link = _db[_ekey].get("telegram_link") or (f"tg://user?id={user_id}")
                _keyboard = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("✅ Принять", callback_data=f"approve_{user_id}"),
                        InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_{user_id}"),
                    ]
                ])
                try:
                    await context.bot.send_message(
                        chat_id=ADMIN_ID,
                        text=(
                            f"📋 <b>Заявка на вступление (уже в базе)</b>\n\n"
                            f"👤 <a href=\"{_tg_link}\">{_u.full_name}</a>\n"
                            f"🏢 Этаж: {_floor}\n"
                            f"🚪 Квартира: {_flat}\n"
                            f"🔗 {_tg}"
                        ),
                        parse_mode="HTML",
                        reply_markup=_keyboard,
                    )
                except Exception:
                    pass
            return

        users[user_id] = {"step": 1}
        try:
            await context.bot.send_message(chat_id=user_id, text="Напиши этаж.")
        except Exception:
            pass

    elif data == "no_house":
        # Очищаем сессию верификации если она была
        users.pop(user_id, None)
        # Убираем кнопки с сообщения
        try:
            await query.edit_message_reply_markup(reply_markup=None)
        except Exception:
            pass
        try:
            await context.bot.send_message(chat_id=user_id, text="Хорошо, если передумаете — напишите боту снова.")
        except Exception:
            pass
        # Уведомляем администратора об отказе
        u = query.from_user
        uname = f"@{u.username}" if u.username else "нет username"
        tg_link = f"https://t.me/{u.username}" if u.username else f"tg://user?id={u.id}"
        try:
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=(
                    f"🚫 <b>Отказался от верификации</b>\n"
                    f"👤 <a href=\"{tg_link}\">{u.full_name}</a>\n"
                    f"Username: {uname}\n"
                    f"ID: <code>{u.id}</code>"
                ),
                parse_mode="HTML",
            )
        except Exception:
            pass

    elif data.startswith("approve_"):
        approved_user_id = int(data.split("_")[1])

        # Принять заявку на вступление в группу
        req_chat_id = join_requests.get(approved_user_id, CHAT_ID)
        try:
            await context.bot.approve_chat_join_request(
                chat_id=req_chat_id, user_id=approved_user_id
            )
        except Exception:
            pass

        # Получить ссылку на чат
        invite_link = None
        try:
            chat = await context.bot.get_chat(req_chat_id)
            invite_link = chat.invite_link
        except Exception:
            pass

        # Отправить финальное сообщение пользователю
        welcome = (
            "✅ Заявка принята, добро пожаловать!\n\n"
            "В чате есть раздел «Этаж, квартира» — "
            "укажите там квартиру и этаж для удобства связи.\n\n"
            "Также доступны рекомендации, барахолка, "
            "полезные контакты и обсуждение вопросов "
            "по управлению домом и заявкам.\n\n"
            "Давайте решать вопросы дома вместе "
            "и создавать приятное соседство."
        )
        if invite_link:
            welcome += f"\n\n🔗 Ссылка на чат: {invite_link}"

        try:
            await context.bot.send_message(chat_id=approved_user_id, text=welcome)
        except Exception:
            pass

        join_requests.pop(approved_user_id, None)
        save_join_requests(join_requests)
        await query.edit_message_text("✅ Заявка одобрена")

    elif data.startswith("reject_"):
        rejected_user_id = int(data.split("_")[1])

        # Отклонить заявку на вступление в группу
        req_chat_id = join_requests.get(rejected_user_id, CHAT_ID)
        try:
            await context.bot.decline_chat_join_request(
                chat_id=req_chat_id, user_id=rejected_user_id
            )
        except Exception:
            pass

        db = load_db()
        if str(rejected_user_id) in db:
            del db[str(rejected_user_id)]
            save_db(db)

        try:
            await context.bot.send_message(
                chat_id=rejected_user_id,
                text="❌ Ваша заявка отклонена."
            )
        except Exception:
            pass

        join_requests.pop(rejected_user_id, None)
        save_join_requests(join_requests)
        await query.edit_message_text("❌ Заявка отклонена")

    elif data.startswith("delconfirm_"):
        flat = int(data.split("_")[1])
        db = load_db()
        keys = [k for k, v in db.items() if v.get("flat") == flat]
        for k in keys:
            del db[k]
        save_db(db)
        await query.edit_message_text(f"✅ Квартира {flat} удалена из базы ({len(keys)} запись).")

    elif data.startswith("delcancel_"):
        flat = int(data.split("_")[1])
        await query.edit_message_text(f"Отмена. Квартира {flat} не удалена.")

    elif data.startswith("resend_"):
        if user_id not in ADMIN_IDS:
            return
        target_id = int(data.split("_", 1)[1])
        try:
            await send_welcome(target_id, context)
            await query.edit_message_reply_markup(reply_markup=query.message.reply_markup)
            await query.answer("✅ Форма отправлена")
        except Exception as e:
            await query.answer(f"❌ {e}", show_alert=True)

    elif data.startswith("flatdel_"):
        if user_id not in ADMIN_IDS:
            return
        key = data[len("flatdel_"):]
        db = load_db()
        if key in db:
            rec = db.pop(key)
            save_db(db)
            name = rec.get("name", "—")
            flat = rec.get("flat", "?")
            await query.edit_message_text(f"🗑 Удалён: {name}, кв. {flat}")
        else:
            await query.edit_message_text("⚠️ Запись уже удалена.")

    elif data.startswith("unreg_dismiss:"):
        uid = data.split(":", 1)[1]
        unreg = load_unregistered()
        name = unreg.pop(uid, {}).get("name", uid)
        save_unregistered(unreg)
        await query.edit_message_text(f"✓ {name} убран из списка незарегистрированных.")

    elif data == "unreg_dismiss_all":
        save_unregistered({})
        await query.edit_message_text("✅ Список незарегистрированных очищен.")

    elif data == "tagflats_confirm":
        await query.answer()
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Да, поставить всем", callback_data="tagflats_do"),
            InlineKeyboardButton("❌ Отмена", callback_data="clearalltags_cancel"),
        ]])
        await query.message.reply_text(
            "🏠 Поставить каждому жильцу тег <b>этаж + квартира</b>?\n"
            "Например: <code>2 эт 348</code>\n\n"
            "Если тег уже есть — не дублируется.",
            parse_mode="HTML",
            reply_markup=keyboard,
        )

    elif data == "tagflats_do":
        db = load_db()
        ok, fail, skipped = 0, 0, 0
        await query.edit_message_text("⏳ Ставлю подписи, подождите...")
        for k, v in db.items():
            floor = v.get("floor", 0)
            flat = v.get("flat", 0)
            uid = v.get("user_id")
            if not flat:
                skipped += 1
                continue
            if v.get("no_tag"):
                skipped += 1
                continue
            tag = f"{floor} эт {flat}" if floor else str(flat)
            # DB тег
            tags = v.get("tags", [])
            if tag not in tags:
                tags.append(tag)
                db[k]["tags"] = tags
            # Telegram-подпись
            if uid:
                success, _ = await _set_member_title(context.bot, int(uid), tag)
                if success:
                    ok += 1
                else:
                    fail += 1
            else:
                ok += 1  # в базе сохранили, Telegram не трогали
        save_db(db)
        await query.edit_message_text(
            f"✅ Подписи поставлены: {ok} жильцов.\n"
            + (f"⚠️ Ошибок: {fail}\n" if fail else "")
            + (f"⏭ Пропущено (нет кв.): {skipped}" if skipped else "")
        )

    elif data == "tagall_prompt":
        await query.answer()
        await query.message.reply_text(
            "Напиши команду чтобы добавить тег всем жильцам:\n"
            "<code>/tagall ТЕГ</code>\n\n"
            "Пример: <code>/tagall жилец</code>",
            parse_mode="HTML",
        )

    elif data == "clearalltags_confirm":
        await query.answer()
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Да, убрать все теги у всех", callback_data="clearalltags_do"),
            InlineKeyboardButton("❌ Отмена", callback_data="clearalltags_cancel"),
        ]])
        await query.message.reply_text(
            "⚠️ Убрать <b>все теги</b> у всех жильцов?\nЭто действие нельзя отменить.",
            parse_mode="HTML",
            reply_markup=keyboard,
        )

    elif data == "clearalltags_do":
        db = load_db()
        await query.edit_message_text("⏳ Снимаю подписи, подождите...")
        db_count, tg_ok, tg_fail = 0, 0, 0
        for k in db:
            # Чистим DB теги
            if db[k].get("tags"):
                db[k]["tags"] = []
                db_count += 1
            # Чистим Telegram-подпись
            uid = db[k].get("user_id")
            if uid:
                success, _ = await _set_member_title(context.bot, int(uid), "")
                if success:
                    tg_ok += 1
                else:
                    tg_fail += 1
        save_db(db)
        await query.edit_message_text(
            f"✅ Telegram-подписи сняты: {tg_ok} жильцов.\n"
            f"🗑 Теги в базе очищены: {db_count} записей.\n"
            + (f"⚠️ Ошибок Telegram: {tg_fail}" if tg_fail else "")
        )

    elif data == "clearalltags_cancel":
        await query.edit_message_text("❌ Отменено.")

    elif data.startswith("settag_"):
        # settag_{user_id}_{floor}_{flat}
        parts = data.split("_")
        try:
            target_uid = int(parts[1])
            floor = int(parts[2])
            flat = int(parts[3])
        except (IndexError, ValueError):
            await query.edit_message_text("⚠️ Неверный формат данных кнопки.")
            return
        tag = f"{floor} эт {flat}" if floor else str(flat)
        tag_ok, err = await _set_member_title(context.bot, target_uid, tag)
        if tag_ok:
            await query.edit_message_text(
                query.message.text_html + f"\n\n✅ <b>Подпись «{tag}» поставлена.</b>",
                parse_mode="HTML",
            )
        else:
            await query.edit_message_text(
                query.message.text_html + f"\n\n⚠️ <b>Не удалось поставить подпись.</b>\n<code>{err}</code>",
                parse_mode="HTML",
            )

    elif data.startswith("linkflat_"):
        if user_id not in ADMIN_IDS:
            return
        parts = data.split("_")
        fwd_uid = int(parts[1])
        flat = int(parts[2])
        db = load_db()
        matched_key = next((k for k, v in db.items() if v.get("flat") == flat), None)
        if matched_key:
            db[matched_key]["user_id"] = fwd_uid
            if db[matched_key].get("telegram", "нет") == "нет":
                db[matched_key]["telegram"] = str(fwd_uid)
                db[matched_key]["telegram_link"] = f"tg://user?id={fwd_uid}"
            save_db(db)
            rec = db[matched_key]
            rec_floor = rec.get("floor", 0)
            tag = f"{rec_floor} эт {flat}" if rec_floor else str(flat)
            pending_forward.pop(user_id, None)
            await query.edit_message_text(
                f"✅ Привязано: {rec.get('name', '—')}, кв. {flat}\n"
                f"ID {fwd_uid} сохранён.\n\n"
                f"Поставить подпись <b>«{tag}»</b>?",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{fwd_uid}_{rec_floor}_{flat}"),
                    InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{fwd_uid}"),
                ]])
            )
        else:
            await query.edit_message_text(f"❌ Квартира {flat} не найдена в базе.")

    elif data.startswith("skiptag_"):
        await query.edit_message_text(
            query.message.text_html + "\n\n❌ <i>Подпись не ставили.</i>",
            parse_mode="HTML",
        )


async def handle_forward(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return

    msg = update.message
    fwd_user_id = None
    fwd_name = None
    fwd_username = None

    # python-telegram-bot v20+ uses forward_origin
    origin = getattr(msg, "forward_origin", None)
    print(f"[forward] origin={origin!r} type={type(origin).__name__}")
    if origin and hasattr(origin, "sender_user") and origin.sender_user:
        fwd_user_id = origin.sender_user.id
        fwd_name = origin.sender_user.full_name
        fwd_username = origin.sender_user.username
        print(f"[forward] sender_user id={fwd_user_id} name={fwd_name} username={fwd_username}")
    elif getattr(msg, "forward_from", None):
        fwd_user_id = msg.forward_from.id
        fwd_name = msg.forward_from.full_name
        fwd_username = msg.forward_from.username
        print(f"[forward] forward_from id={fwd_user_id} name={fwd_name} username={fwd_username}")
    else:
        # Попытка достать имя из скрытого origin
        hidden_name = None
        if origin and hasattr(origin, "sender_user_name"):
            hidden_name = origin.sender_user_name
        print(f"[forward] privacy on, hidden_name={hidden_name}")

    if not fwd_user_id:
        await msg.reply_text(
            "⚠️ Закрытый профиль — ID получить нельзя.\n\n"
            "Используй команду:\n"
            "<code>/setid @username ID</code>\n\n"
            "Например: <code>/setid @Squalla 123456789</code>\n\n"
            "ID можно узнать переслав сообщение боту @userinfobot",
            parse_mode="HTML",
        )
        return

    db = load_db()

    # Уже привязан?
    for key, v in db.items():
        if str(v.get("user_id")) == str(fwd_user_id) or key == str(fwd_user_id):
            # Если username раньше был «нет» — обновляем
            old_tg = str(v.get("telegram", ""))
            upd_note = ""
            if fwd_username and (old_tg in ("", "нет") or not old_tg.startswith("@")):
                db[key]["telegram"] = f"@{fwd_username}"
                db[key]["telegram_link"] = f"https://t.me/{fwd_username}"
                save_db(db)
                upd_note = f"\n🔄 Username обновлён: @{fwd_username}"
            rec_floor = v.get("floor", 0)
            rec_flat = v.get("flat", 0)
            tag = f"{rec_floor} эт {rec_flat}" if rec_floor else str(rec_flat)
            keyboard = None
            if rec_flat:
                keyboard = InlineKeyboardMarkup([[
                    InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{fwd_user_id}_{rec_floor}_{rec_flat}"),
                    InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{fwd_user_id}"),
                ]])
            await msg.reply_text(
                f"✅ Уже в базе: {v['name']}, кв. {rec_flat}{upd_note}\n\n"
                + (f"Поставить подпись <b>«{tag}»</b>?" if rec_flat else ""),
                parse_mode="HTML",
                reply_markup=keyboard,
            )
            return

    # Получаем @username из пересланного сообщения
    fwd_username = None
    if origin and hasattr(origin, "sender_user") and origin.sender_user:
        fwd_username = origin.sender_user.username
    elif getattr(msg, "forward_from", None):
        fwd_username = msg.forward_from.username

    # Автоматически сопоставляем по @username с базой
    if fwd_username:
        uname_lower = f"@{fwd_username}".lower()
        for key, v in db.items():
            if v.get("user_id"):
                continue
            tg = str(v.get("telegram", "")).lower()
            tg_link = str(v.get("telegram_link", "")).lower()
            rec_uname = None
            if tg.startswith("@"):
                rec_uname = tg
            elif tg_link.startswith("https://t.me/"):
                rec_uname = "@" + tg_link.replace("https://t.me/", "").rstrip("/")
            if rec_uname and rec_uname == uname_lower:
                db[key]["user_id"] = fwd_user_id
                save_db(db)
                flat = v.get("flat", 0)
                floor = v.get("floor", 0)
                tag = f"{floor} эт {flat}" if floor else str(flat)
                tag_ok = False
                if flat:
                    tag_ok, _ = await _set_member_title(context.bot, fwd_user_id, tag)
                tag_info = f"🏷 Подпись «{tag}» поставлена" if tag_ok else "⚠️ Подпись не удалось поставить автоматически"
                await msg.reply_text(
                    f"✅ Автоматически привязан!\n"
                    f"👤 {v['name']}, кв. {flat}\n"
                    f"ID {fwd_user_id} сохранён.\n"
                    f"{tag_info}"
                )
                return

    # Не нашли в базе — пробуем распознать квартиру из текста
    admin_uid = update.effective_user.id
    pending_forward[admin_uid] = {"user_id": fwd_user_id, "name": fwd_name, "username": fwd_username}
    uname_str = f"@{fwd_username}" if fwd_username else "нет username"
    fwd_text = msg.text or getattr(msg, "caption", None) or ""
    parsed_floor, parsed_flat = parse_floor_flat(fwd_text)
    if parsed_flat:
        matched_key = next((k for k, v in db.items() if v.get("flat") == parsed_flat), None)
        if matched_key:
            res_name = db[matched_key].get("name", "—")
            floor_hint = f"этаж {parsed_floor}, " if parsed_floor else ""
            await msg.reply_text(
                f"👤 {fwd_name} ({uname_str})\n"
                f"🆔 ID: {fwd_user_id}\n\n"
                f"📝 В тексте найдено: {floor_hint}квартира <b>{parsed_flat}</b>\n"
                f"🏠 В базе: {res_name}\n\n"
                f"Подтвердить привязку?",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(f"✅ Привязать кв. {parsed_flat}", callback_data=f"linkflat_{fwd_user_id}_{parsed_flat}"),
                    InlineKeyboardButton("✏️ Другой номер", callback_data=f"skiptag_{fwd_user_id}"),
                ]])
            )
            return
    await msg.reply_text(
        f"👤 {fwd_name} ({uname_str})\n"
        f"🆔 ID: {fwd_user_id}\n\n"
        f"Не найден в базе. Напиши номер квартиры для привязки (или /cancel):"
    )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Игнорируем всё что не личное сообщение боту
    if not update.message or update.message.chat.type != "private":
        return

    user_id = update.effective_user.id

    # Admin: waiting to enter flat number after forwarded message
    if user_id in ADMIN_IDS and user_id in pending_forward:
        text = update.message.text.strip()
        if text.isdigit():
            flat = int(text)
            db = load_db()
            matched_key = None
            for key, v in db.items():
                if v.get("flat") == flat:
                    matched_key = key
                    break

            info = pending_forward.pop(user_id)
            fwd_uid = info["user_id"]
            fwd_name = info["name"]

            if matched_key:
                db[matched_key]["user_id"] = fwd_uid
                if db[matched_key].get("telegram", "нет") == "нет":
                    db[matched_key]["telegram"] = str(fwd_uid)
                    db[matched_key]["telegram_link"] = f"tg://user?id={fwd_uid}"
                save_db(db)
                rec_floor = db[matched_key].get("floor", 0)
                rec_flat = db[matched_key].get("flat", flat)
                tag = f"{rec_floor} эт {rec_flat}" if rec_floor else str(rec_flat)
                keyboard = None
                if rec_flat:
                    keyboard = InlineKeyboardMarkup([[
                        InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{fwd_uid}_{rec_floor}_{rec_flat}"),
                        InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{fwd_uid}"),
                    ]])
                await update.message.reply_text(
                    f"✅ Привязано: {fwd_name} → кв. {rec_flat} ({db[matched_key]['name']})\n"
                    f"ID {fwd_uid} сохранён.\n\n"
                    + (f"Поставить подпись <b>«{tag}»</b>?" if rec_flat else ""),
                    parse_mode="HTML",
                    reply_markup=keyboard,
                )
            else:
                # Квартиры нет в базе — проверяем что она вообще существует в доме
                rec_floor = next((fl for fl, flats in floor_flats.items() if flat in flats), None)
                if rec_floor:
                    # Валидная квартира — создаём новую запись
                    fwd_username_saved = info.get("username")
                    tg_str = f"@{fwd_username_saved}" if fwd_username_saved else str(fwd_uid)
                    tg_link = f"https://t.me/{fwd_username_saved}" if fwd_username_saved else f"tg://user?id={fwd_uid}"
                    new_key = str(fwd_uid)
                    db[new_key] = {
                        "name": fwd_name,
                        "floor": rec_floor,
                        "flat": flat,
                        "telegram": tg_str,
                        "telegram_link": tg_link,
                        "user_id": fwd_uid,
                        "date_added": datetime.date.today().isoformat(),
                    }
                    save_db(db)
                    tag = f"{rec_floor} эт {flat}"
                    keyboard = InlineKeyboardMarkup([[
                        InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{fwd_uid}_{rec_floor}_{flat}"),
                        InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{fwd_uid}"),
                    ]])
                    await update.message.reply_text(
                        f"✅ Создана новая запись: {fwd_name} → кв. {flat} ({rec_floor} эт.)\n"
                        f"ID {fwd_uid} сохранён.\n\n"
                        f"Поставить подпись <b>«{tag}»</b>?",
                        parse_mode="HTML",
                        reply_markup=keyboard,
                    )
                else:
                    await update.message.reply_text(
                        f"❌ Квартира {flat} не существует в этом доме (диапазон 312–531)."
                    )
            return

    if user_id not in users:
        # Если пишут боту в личку не в процессе верификации — показываем приветствие заново
        if update.message.chat.type == "private" and user_id not in ADMIN_IDS:
            try:
                await send_welcome(user_id, context)
            except Exception:
                pass
        return

    text = update.message.text
    step = users[user_id]["step"]

    if step == 1:
        try:
            floor = int(text)

            if floor < 2 or floor > 23:
                await update.message.reply_text("❌ Этаж должен быть от 2 до 23.")
                return

            users[user_id]["floor"] = floor
            users[user_id]["step"] = 2
            await update.message.reply_text("Теперь напиши номер квартиры.")

        except ValueError:
            await update.message.reply_text("Введите этаж числом.")

    elif step == 2:
        try:
            flat = int(text)
            floor = users[user_id]["floor"]

            if flat not in floor_flats[floor]:
                await update.message.reply_text(
                    "❌ Проверьте правильность ввода квартиры по ДДУ."
                )
                return

            users[user_id]["flat"] = flat

            tg_username = update.effective_user.username
            if tg_username:
                tg_link = f"https://t.me/{tg_username}"
                tg_text = f"@{tg_username}"
            else:
                tg_link = "нет"
                tg_text = "нет"

            full_name = update.effective_user.full_name

            residents = load_db()

            # Удалить дубли: по имени+квартире, по user_id или по telegram username
            duplicates = [
                k for k, v in residents.items()
                if (
                    (v.get("flat") == flat and v.get("floor") == floor and v.get("name") == full_name)
                    or v.get("user_id") == user_id
                    or (tg_text not in ("нет", "") and v.get("telegram", "").lower() == tg_text.lower())
                )
            ]
            for k in duplicates:
                del residents[k]

            new_key = str(user_id)
            universal_tags = get_universal_tags(residents, exclude_key=new_key)
            residents[new_key] = {
                "name": full_name,
                "floor": floor,
                "flat": flat,
                "telegram": tg_text,
                "telegram_link": tg_link,
                "user_id": user_id,
                "date_added": datetime.date.today().isoformat(),
            }
            if universal_tags:
                residents[new_key]["tags"] = universal_tags
            save_db(residents)

            user_link = tg_link if tg_link != "нет" else f"tg://user?id={user_id}"
            link_display = f"@{tg_username}" if tg_username else f"tg://user?id={user_id}"

            if user_id in join_requests:
                # Человек подаёт заявку на вступление — нужно одобрение
                keyboard = InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton("✅ Принять", callback_data=f"approve_{user_id}"),
                        InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_{user_id}"),
                    ]
                ])
                admin_text = (
                    f"📋 <b>Новая заявка на вступление</b>\n\n"
                    f"👤 <a href=\"{user_link}\">{full_name}</a>\n"
                    f"🏢 Этаж: {floor}\n"
                    f"🚪 Квартира: {flat}\n"
                    f"🔗 {link_display}"
                )
                await context.bot.send_message(
                    chat_id=ADMIN_ID,
                    text=admin_text,
                    parse_mode="HTML",
                    reply_markup=keyboard,
                )
            else:
                # Человек уже в чате, просто верифицировался
                admin_text = (
                    f"✅ <b>Новый жилец записан в базу</b>\n\n"
                    f"👤 <a href=\"{user_link}\">{full_name}</a>\n"
                    f"🏢 Этаж: {floor}\n"
                    f"🚪 Квартира: {flat}\n"
                    f"🔗 {link_display}"
                )
                await context.bot.send_message(
                    chat_id=ADMIN_ID,
                    text=admin_text,
                    parse_mode="HTML",
                )

            if user_id in join_requests:
                await update.message.reply_text(
                    "Спасибо за информацию! 🙏\n"
                    "Ваша заявка передана администратору — скоро откроем доступ в чат."
                )
            else:
                await update.message.reply_text("Спасибо за информацию! 🙏")
                # Уже в чате — спрашиваем администратора про тег
                tag = f"{floor} эт {flat}"
                keyboard = InlineKeyboardMarkup([[
                    InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{user_id}_{floor}_{flat}"),
                    InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{user_id}"),
                ]])
                try:
                    await context.bot.send_message(
                        chat_id=ADMIN_ID,
                        text=f"Поставить Telegram-подпись <b>«{tag}»</b>?",
                        parse_mode="HTML",
                        reply_markup=keyboard,
                    )
                except Exception:
                    pass
            del users[user_id]

        except ValueError:
            await update.message.reply_text("Введите квартиру числом.")


def _u16(s: str) -> int:
    return len(s.encode("utf-16-le")) // 2


def _make_entity(type_, offset, length, url=None, user=None):
    kwargs = {"type": type_, "offset": offset, "length": length}
    if url:
        kwargs["url"] = url
    if user:
        kwargs["user"] = user
    return MessageEntity(**kwargs)


async def residents_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return

    residents = load_db()

    if not residents:
        await update.message.reply_text("База пустая.")
        return

    floors = {}
    for key, data in residents.items():
        floor = data["floor"]
        if floor not in floors:
            floors[floor] = []
        entry = dict(data)
        entry["_key"] = key
        floors[floor].append(entry)

    messages = []

    header = "📋 Жильцы:\n\n"
    cur_text = header
    cur_entities = [_make_entity(MessageEntity.BOLD, _u16("📋 "), _u16("Жильцы:"))]

    for floor in sorted(floors.keys()):
        block = f"🏢 Этаж {floor}\n"
        block_ents = [_make_entity(MessageEntity.BOLD, 0, _u16(block.rstrip("\n")))]

        # Группируем жильцов по квартирам
        flat_groups = {}
        for resident in floors[floor]:
            flat = resident["flat"]
            if flat not in flat_groups:
                flat_groups[flat] = []
            flat_groups[flat].append(resident)

        for flat in sorted(flat_groups.keys()):
            group = flat_groups[flat]
            prefix = f"• кв. {flat} — "
            # Имена для отображения: добавляем ❌ если жилец покинул чат
            display_names = [(r["name"] + " ❌") if r.get("left") else r["name"] for r in group]
            # Собираем теги со всех жильцов квартиры
            flat_tags: set = set()
            for r in group:
                flat_tags.update(r.get("tags", []))
            tag_str = f"  [{', '.join(sorted(flat_tags))}]" if flat_tags else ""
            line = prefix + ", ".join(display_names) + tag_str + "\n"

            # Смещение начала первого имени внутри block
            name_cursor = _u16(block) + _u16(prefix)

            for i, resident in enumerate(group):
                name = resident["name"]
                name_len = _u16(name)  # ссылка покрывает только имя, без ❌
                display_name_len = _u16(display_names[i])
                key = resident["_key"]
                tg_link = resident.get("telegram_link", "")
                tg = resident.get("telegram", "нет")
                uid = resident.get("user_id")

                entity = None
                if tg and tg != "нет" and tg.startswith("@"):
                    entity = _make_entity(MessageEntity.TEXT_LINK, name_cursor, name_len,
                                          url=f"https://t.me/{tg[1:]}")
                elif tg_link and tg_link != "нет" and tg_link.startswith("https://"):
                    entity = _make_entity(MessageEntity.TEXT_LINK, name_cursor, name_len,
                                          url=tg_link)
                elif uid:
                    tg_user = TGUser(id=int(uid), first_name=name, is_bot=False)
                    entity = _make_entity(MessageEntity.TEXT_MENTION, name_cursor, name_len,
                                          user=tg_user)
                elif tg and tg != "нет" and tg.lstrip("-").isdigit():
                    tg_user = TGUser(id=int(tg), first_name=name, is_bot=False)
                    entity = _make_entity(MessageEntity.TEXT_MENTION, name_cursor, name_len,
                                          user=tg_user)
                elif key.lstrip("-").isdigit():
                    tg_user = TGUser(id=int(key), first_name=name, is_bot=False)
                    entity = _make_entity(MessageEntity.TEXT_MENTION, name_cursor, name_len,
                                          user=tg_user)

                if entity:
                    block_ents.append(entity)

                # Сдвигаем курсор: отображаемое имя (с ❌ если есть) + ", " (кроме последнего)
                if i < len(group) - 1:
                    name_cursor += display_name_len + _u16(", ")

            block += line

        block += "\n"

        if _u16(cur_text) + _u16(block) > 4000:
            messages.append((cur_text, cur_entities))
            cur_text = block
            cur_entities = list(block_ents)
        else:
            adj = _u16(cur_text)
            for e in block_ents:
                cur_entities.append(_make_entity(
                    e.type, e.offset + adj, e.length,
                    url=getattr(e, "url", None),
                    user=getattr(e, "user", None)
                ))
            cur_text += block

    if cur_text:
        messages.append((cur_text, cur_entities))

    for text, entities in messages:
        await update.message.reply_text(text, entities=entities)


async def tag_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Добавить тег жильцу: /tag 348 председатель"""
    if update.effective_user.id not in ADMIN_IDS:
        return

    if len(context.args) < 2:
        await update.message.reply_text(
            "Использование: <code>/tag номер_квартиры тег</code>\n"
            "Пример: <code>/tag 348 председатель</code>",
            parse_mode="HTML"
        )
        return

    try:
        flat = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Номер квартиры должен быть числом.")
        return

    tag = " ".join(context.args[1:]).strip().lower()
    db = load_db()
    keys = [k for k, v in db.items() if v.get("flat") == flat]

    if not keys:
        await update.message.reply_text(f"❌ Квартира {flat} не найдена в базе.")
        return

    updated = []
    for k in keys:
        tags = db[k].get("tags", [])
        if tag not in tags:
            tags.append(tag)
            db[k]["tags"] = tags
            updated.append(db[k]["name"])

    save_db(db)
    if updated:
        await update.message.reply_text(
            f"✅ Тег <b>{tag}</b> добавлен к кв. {flat}: {', '.join(updated)}",
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(f"ℹ️ Тег <b>{tag}</b> уже есть у кв. {flat}.", parse_mode="HTML")


async def untag_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Удалить тег у жильца: /untag 348 председатель"""
    if update.effective_user.id not in ADMIN_IDS:
        return

    if len(context.args) < 2:
        await update.message.reply_text(
            "Использование: <code>/untag номер_квартиры тег</code>\n"
            "Пример: <code>/untag 348 председатель</code>",
            parse_mode="HTML"
        )
        return

    try:
        flat = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Номер квартиры должен быть числом.")
        return

    tag = " ".join(context.args[1:]).strip().lower()
    db = load_db()
    keys = [k for k, v in db.items() if v.get("flat") == flat]

    if not keys:
        await update.message.reply_text(f"❌ Квартира {flat} не найдена в базе.")
        return

    removed = []
    for k in keys:
        tags = db[k].get("tags", [])
        if tag in tags:
            tags.remove(tag)
            db[k]["tags"] = tags
            removed.append(db[k]["name"])

    save_db(db)
    if removed:
        await update.message.reply_text(
            f"✅ Тег <b>{tag}</b> удалён у кв. {flat}: {', '.join(removed)}",
            parse_mode="HTML"
        )
    else:
        await update.message.reply_text(f"ℹ️ Тега <b>{tag}</b> нет у кв. {flat}.", parse_mode="HTML")


async def tags_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Список всех тегов в базе: /tags"""
    if update.effective_user.id not in ADMIN_IDS:
        return

    db = load_db()
    tag_map: dict[str, list[str]] = {}
    for v in db.values():
        for t in v.get("tags", []):
            tag_map.setdefault(t, []).append(f"кв. {v.get('flat')} {v.get('name', '')}")

    # Кнопки массового управления тегами
    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🏠 Поставить этаж+кв всем", callback_data="tagflats_confirm"),
            InlineKeyboardButton("🗑 Убрать все теги у всех", callback_data="clearalltags_confirm"),
        ],
        [
            InlineKeyboardButton("🏷 Добавить одинаковый тег всем", callback_data="tagall_prompt"),
        ],
    ])

    if not tag_map:
        await update.message.reply_text(
            "🏷 Тегов пока нет.\n\n"
            "Добавить тег всем: <code>/tagall ТЕГ</code>\n"
            "Убрать тег у всех: <code>/untagall ТЕГ</code>",
            parse_mode="HTML",
            reply_markup=keyboard,
        )
        return

    lines = ["🏷 <b>Все теги:</b>\n"]
    for tag in sorted(tag_map.keys()):
        residents_str = ", ".join(tag_map[tag])
        lines.append(f"<b>{tag}</b> ({len(tag_map[tag])}) — {residents_str}")

    lines.append("\n<i>Добавить тег всем: /tagall ТЕГ\nУбрать тег у всех: /untagall ТЕГ</i>")
    await update.message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=keyboard)


async def tagall_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/tagall ТЕГ — добавить тег всем жильцам в базе."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    if not context.args:
        await update.message.reply_text(
            "Использование: <code>/tagall ТЕГ</code>\nПример: <code>/tagall жилец</code>",
            parse_mode="HTML",
        )
        return
    tag = " ".join(context.args).strip().lower()
    db = load_db()
    count = 0
    for k in db:
        tags = db[k].get("tags", [])
        if tag not in tags:
            tags.append(tag)
            db[k]["tags"] = tags
            count += 1
    save_db(db)
    await update.message.reply_text(
        f"✅ Тег <b>{tag}</b> добавлен {count} жильцам.",
        parse_mode="HTML",
    )


async def untagall_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/untagall ТЕГ — убрать тег у всех жильцов."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    if not context.args:
        await update.message.reply_text(
            "Использование: <code>/untagall ТЕГ</code>\nПример: <code>/untagall жилец</code>",
            parse_mode="HTML",
        )
        return
    tag = " ".join(context.args).strip().lower()
    db = load_db()
    count = 0
    for k in db:
        tags = db[k].get("tags", [])
        if tag in tags:
            tags.remove(tag)
            db[k]["tags"] = tags
            count += 1
    save_db(db)
    await update.message.reply_text(
        f"✅ Тег <b>{tag}</b> убран у {count} жильцов.",
        parse_mode="HTML",
    )


async def _tg_api(method: str, **kwargs) -> dict:
    """Прямой вызов Telegram Bot API (для методов не реализованных в PTB)."""
    import httpx
    token = os.environ["TELEGRAM_BOT_TOKEN"]
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"https://api.telegram.org/bot{token}/{method}",
            json=kwargs,
            timeout=10,
        )
        return resp.json()


async def _set_member_title(bot, user_id: int, title: str) -> tuple[bool, str]:
    """Ставит тег участнику группы через setChatMemberTag.
    Возвращает (True, "") при успехе или (False, описание_ошибки) при неудаче.
    Не работает для создателя чата (CHAT_CREATOR_REQUIRED) — это нормально, пропускаем.
    При flood wait — ждёт и повторяет автоматически.
    Пропускает жильцов с флагом no_tag: True (только при установке, не при очистке)."""
    if title:  # проверяем только когда ставим тег, не когда убираем
        db = load_db()
        for v in db.values():
            if v.get("user_id") == user_id and v.get("no_tag"):
                return False, "no_tag"
    for attempt in range(4):
        result = await _tg_api("setChatMemberTag", chat_id=CHAT_ID, user_id=user_id, tag=title)
        if result.get("ok"):
            return True, ""
        err = result.get("description", str(result))
        retry_after = result.get("parameters", {}).get("retry_after", 0)
        if retry_after and attempt < 3:
            await asyncio.sleep(retry_after + 1)
            continue
        return False, err
    return False, "max retries exceeded"


async def _resolve_uid(context, r: dict, db_key: str | None = None) -> int | None:
    """Пытается получить числовой user_id из записи жильца. Сохраняет в базу при нахождении."""
    # 1. Прямой user_id
    if r.get("user_id"):
        return int(r["user_id"])

    tg = str(r.get("telegram", ""))
    tg_link = str(r.get("telegram_link", ""))

    # 2. telegram — числовая строка (это и есть user_id)
    if tg and tg.lstrip("-").isdigit():
        uid = int(tg)
        if db_key:
            db = load_db()
            if db_key in db:
                db[db_key]["user_id"] = uid
                save_db(db)
        return uid

    # 3. telegram_link = tg://user?id=12345
    if tg_link.startswith("tg://user?id="):
        try:
            uid = int(tg_link.split("=", 1)[1])
            if db_key:
                db = load_db()
                if db_key in db:
                    db[db_key]["user_id"] = uid
                    save_db(db)
            return uid
        except Exception:
            pass

    # 4. @username — пробуем getChatMember в контексте группы (знает всех участников)
    username = None
    if tg.startswith("@"):
        username = tg
    elif tg_link.startswith("https://t.me/"):
        username = "@" + tg_link.replace("https://t.me/", "").rstrip("/")

    if username:
        # Сначала через getChatMember — надёжнее для участников группы
        try:
            result = await _tg_api("getChatMember", chat_id=CHAT_ID, user_id=username)
            if result.get("ok"):
                uid = result["result"]["user"]["id"]
                if db_key:
                    db = load_db()
                    if db_key in db:
                        db[db_key]["user_id"] = uid
                        save_db(db)
                return uid
        except Exception:
            pass
        # Fallback: getChat
        try:
            chat = await context.bot.get_chat(username)
            uid = chat.id
            if db_key and uid:
                db = load_db()
                if db_key in db:
                    db[db_key]["user_id"] = uid
                    save_db(db)
            return uid
        except Exception:
            pass

    return None


async def settitles_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Проставляет подписи 'Кв. XXX' всем жильцам с user_id в базе."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    db = load_db()
    candidates = [
        (k, v) for k, v in db.items()
        if v.get("flat") and (
            v.get("user_id") or
            str(v.get("telegram", "")).lstrip("-").isdigit() or
            str(v.get("telegram", "")).startswith("@") or
            str(v.get("telegram_link", "")).startswith("tg://user?id=") or
            str(v.get("telegram_link", "")).startswith("https://t.me/")
        )
    ]

    if not candidates:
        await update.message.reply_text("❌ Нет жильцов с Telegram ID или @username.")
        return

    total = len(candidates)
    progress_msg = await update.message.reply_text(f"⏳ Обрабатываю 0 / {total}...")

    ok, fail, skip, admins = [], [], [], []
    for i, (k, r) in enumerate(candidates, 1):
        # Прогресс каждые 10 человек
        if i % 10 == 1 or i == 1:
            try:
                await progress_msg.edit_text(f"⏳ Обрабатываю {i} / {total}...")
            except Exception:
                pass

        uid = await _resolve_uid(context, r, db_key=k)
        if not uid:
            fail.append(f"{r['name']}: не удалось определить ID")
            await asyncio.sleep(0.3)
            continue

        flat = r["flat"]
        floor = r.get("floor", 0)
        tag = f"{floor} эт {flat}" if floor else f"{flat}"

        success, err = await _set_member_title(context.bot, uid, tag)
        if success:
            ok.append(f"{r['name']} → {tag}")
        else:
            if "not a member" in err.lower() or "USER_NOT_PARTICIPANT" in err:
                skip.append(r["name"])
            elif "CHAT_CREATOR_REQUIRED" in err:
                admins.append(r["name"])  # владелец чата — пропускаем
            else:
                fail.append(f"{r['name']}: {err}")

        await asyncio.sleep(1.5)  # пауза между запросами

    try:
        await progress_msg.edit_text(f"✅ Готово!")
    except Exception:
        pass

    lines = [f"✅ Готово — обновлено {len(ok)} из {total}"]
    if ok:
        lines.append("\n<b>Проставлено:</b>\n" + "\n".join(f"• {x}" for x in ok[:30]))
    if skip:
        lines.append(f"\n<b>Не в группе ({len(skip)}):</b> " + ", ".join(skip[:15]))
    if admins:
        lines.append(f"\n<b>Администраторы — тег вручную ({len(admins)}):</b> " + ", ".join(admins))

    await update.message.reply_text("\n".join(lines), parse_mode="HTML")

    # Ошибки отдельным сообщением чтобы не обрезало
    if fail:
        err_lines = [f"<b>Ошибки ({len(fail)}):</b>"]
        err_lines += [f"• {x}" for x in fail[:20]]
        await update.message.reply_text("\n".join(err_lines), parse_mode="HTML")


async def debugtitle_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Диагностика: проверяет права бота и пробует поставить тег одному жильцу."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    lines = []

    # 1. Проверяем права самого бота в группе
    try:
        me = await context.bot.get_me()
        member = await context.bot.get_chat_member(CHAT_ID, me.id)
        lines.append(f"🤖 <b>Бот:</b> {me.username}, статус: {member.status}")
        if hasattr(member, "can_promote_members"):
            lines.append(f"can_promote_members: {member.can_promote_members}")
        if hasattr(member, "can_manage_chat"):
            lines.append(f"can_manage_chat: {member.can_manage_chat}")
    except Exception as e:
        lines.append(f"❌ Ошибка получения прав бота: {e}")

    # 2. Берём первого жильца с ID и пробуем поставить тег
    db = load_db()
    test_resident = None
    for v in db.values():
        if v.get("user_id") and v.get("flat"):
            test_resident = v
            break

    if not test_resident:
        lines.append("\n⚠️ Нет жильцов с user_id для теста")
    else:
        uid = int(test_resident["user_id"])
        flat = test_resident["flat"]
        name = test_resident["name"]
        lines.append(f"\n🧪 <b>Тест на:</b> {name}, кв.{flat}, uid={uid}")

        # Проверяем статус пользователя в группе
        try:
            cm = await context.bot.get_chat_member(CHAT_ID, uid)
            lines.append(f"Статус в группе: {cm.status}")
        except Exception as e:
            lines.append(f"get_chat_member ошибка: {e}")

        # Пробуем _set_member_title
        tag_test = f"{r.get('floor', 0)} эт {flat}" if r.get("floor") else str(flat)
        ok_title, err_title = await _set_member_title(context.bot, uid, tag_test)
        if ok_title:
            lines.append(f"✅ Подпись установлена: {tag_test}")
        else:
            lines.append(f"❌ Подпись: {err_title}")

    await update.message.reply_text("\n".join(lines), parse_mode="HTML")


async def cleartitles_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Снимает подписи и убирает статус псевдо-администратора у всех жильцов."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    db = load_db()
    residents_with_id = [v for v in db.values() if v.get("user_id")]
    if not residents_with_id:
        await update.message.reply_text("❌ Нет жильцов с привязанным Telegram ID.")
        return

    await update.message.reply_text(f"⏳ Снимаю подписи у {len(residents_with_id)} жильцов...")

    ok, fail = 0, 0
    for r in residents_with_id:
        uid = int(r["user_id"])
        try:
            await context.bot.promote_chat_member(
                chat_id=CHAT_ID,
                user_id=uid,
                can_change_info=False,
                can_delete_messages=False,
                can_invite_users=False,
                can_restrict_members=False,
                can_pin_messages=False,
                can_promote_members=False,
                can_manage_chat=False,
                can_manage_video_chats=False,
            )
            ok += 1
        except Exception:
            fail += 1

    await update.message.reply_text(
        f"✅ Подписи сняты: {ok} жильцов.\n"
        + (f"⚠️ Ошибок: {fail}" if fail else "")
    )


async def add_user_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return

    args = context.args
    if not args or len(args) < 3:
        await update.message.reply_text(
            "Использование:\n"
            "<code>/adduser @username этаж квартира Имя Фамилия</code>\n\n"
            "Если нет username:\n"
            "<code>/adduser нет этаж квартира Имя Фамилия</code>\n\n"
            "Пример:\n"
            "<code>/adduser @ivanov 5 345 Иван Иванов</code>",
            parse_mode="HTML"
        )
        return

    username_raw = args[0]
    try:
        floor = int(args[1])
        flat = int(args[2])
    except ValueError:
        await update.message.reply_text("❌ Этаж и квартира должны быть числами.")
        return

    name = " ".join(args[3:]) if len(args) > 3 else "Не указано"

    if floor < 2 or floor > 23:
        await update.message.reply_text("❌ Этаж должен быть от 2 до 23.")
        return

    if flat not in floor_flats.get(floor, range(0)):
        await update.message.reply_text("❌ Такая квартира не соответствует этому этажу.")
        return

    if username_raw.lower() == "нет" or username_raw == "-":
        tg_text = "нет"
        tg_link = "нет"
        key = f"manual_{floor}_{flat}"
    else:
        username = username_raw.lstrip("@")
        tg_text = f"@{username}"
        tg_link = f"https://t.me/{username}"
        key = f"manual_{username}"

    residents = load_db()
    universal_tags = get_universal_tags(residents, exclude_key=key)
    new_record: dict = {
        "name": name,
        "floor": floor,
        "flat": flat,
        "telegram": tg_text,
        "telegram_link": tg_link
    }
    if universal_tags:
        new_record["tags"] = universal_tags
    residents[key] = new_record
    save_db(residents)

    await update.message.reply_text(
        f"✅ Жилец добавлен:\n"
        f"👤 {name}\n"
        f"🏢 Этаж: {floor}\n"
        f"🚪 Квартира: {flat}\n"
        f"🔗 Telegram: {tg_text}"
    )


async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return

    db = load_db()

    # Все квартиры в доме
    all_flats = set()
    for flats in floor_flats.values():
        all_flats.update(flats)
    total_flats = len(all_flats)

    # Уникальные квартиры в базе
    registered_flats = set(v.get("flat") for v in db.values() if v.get("flat"))
    registered_count = len(registered_flats)

    # С Telegram
    with_tg = sum(
        1 for v in db.values()
        if v.get("telegram") and v.get("telegram") not in ("нет", "")
    )

    # Прогресс бар
    pct = registered_count / total_flats * 100
    filled = int(pct / 5)
    bar = "█" * filled + "░" * (20 - filled)

    lines = [
        "📊 <b>Статистика дома 622</b>\n",
        f"🏢 Всего квартир: <b>{total_flats}</b>",
        f"✅ В базе: <b>{registered_count}</b>",
        f"💬 С Telegram: <b>{with_tg}</b>",
        f"❌ Не зарегистрированы: <b>{total_flats - registered_count}</b>",
        f"\n<code>{bar}</code> {pct:.0f}%\n",
        "📐 <b>По этажам:</b>",
    ]

    for floor in range(2, 24):
        flats_on_floor = set(floor_flats[floor])
        total_on_floor = len(flats_on_floor)
        done_on_floor = len(flats_on_floor & registered_flats)
        bar_f = "●" * done_on_floor + "○" * (total_on_floor - done_on_floor)
        lines.append(f"  {floor:>2} эт: {bar_f} {done_on_floor}/{total_on_floor}")

    await update.message.reply_text("\n".join(lines), parse_mode="HTML")


def _build_stats_text() -> str:
    """Строит текст статистики по квартирам и этажам."""
    db = load_db()
    all_flats = set()
    for flats in floor_flats.values():
        all_flats.update(flats)
    total_flats = len(all_flats)
    registered_flats = set(v.get("flat") for v in db.values() if v.get("flat"))
    registered_count = len(registered_flats)
    missing = total_flats - registered_count
    pct = registered_count / total_flats * 100
    filled = int(pct / 5)
    bar = "█" * filled + "░" * (20 - filled)
    floor_lines = []
    for floor in range(2, 24):
        flats_on_floor = set(floor_flats[floor])
        total_on_floor = len(flats_on_floor)
        done_on_floor = len(flats_on_floor & registered_flats)
        dots = "●" * done_on_floor + "○" * (total_on_floor - done_on_floor)
        floor_lines.append(f"  {floor:>2} эт: {dots} {done_on_floor}/{total_on_floor}")
    return (
        "👋 Привет, соседи!\n\n"
        f"В нашем доме <b>{total_flats}</b> квартир (312–531), "
        f"и уже <b>{registered_count}</b> из них зарегистрированы в чате.\n\n"
        f"<code>{bar}</code> {pct:.0f}%\n\n"
        f"Ещё <b>{missing}</b> квартир пока не с нами — "
        "позовите соседей! 🏠\n\n"
        "📐 <b>По этажам:</b>\n"
        + "\n".join(floor_lines)
    )


async def _post_stats(bot) -> int | None:
    """Удаляет предыдущее сообщение статистики и публикует новое. Возвращает ID нового сообщения."""
    # Удаляем предыдущее
    prev_id = load_last_stats_msg()
    if prev_id and STATS_THREAD_ID:
        try:
            await bot.delete_message(chat_id=CHAT_ID, message_id=prev_id)
        except Exception:
            pass

    kwargs = dict(chat_id=CHAT_ID, text=_build_stats_text(), parse_mode="HTML")
    if STATS_THREAD_ID:
        kwargs["message_thread_id"] = STATS_THREAD_ID

    msg = await bot.send_message(**kwargs)
    save_last_stats_msg(msg.message_id)
    return msg.message_id


async def poststats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return
    try:
        await _post_stats(context.bot)
        await update.message.reply_text("✅ Статистика отправлена в чат.")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


LAST_DAILY_RUN_FILE = os.path.join(os.path.dirname(__file__), "last_daily_run.json")


def load_last_daily_run() -> dict:
    if os.path.exists(LAST_DAILY_RUN_FILE):
        try:
            with open(LAST_DAILY_RUN_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_last_daily_run(data: dict):
    with open(LAST_DAILY_RUN_FILE, "w") as f:
        json.dump(data, f)


async def daily_stats(context: ContextTypes.DEFAULT_TYPE):
    """Ежедневная публикация статистики утром."""
    try:
        await _post_stats(context.bot)
        runs = load_last_daily_run()
        runs["stats"] = datetime.date.today().isoformat()
        save_last_daily_run(runs)
    except Exception as e:
        print(f"[daily_stats] Ошибка: {e}")


async def daily_verify(context: ContextTypes.DEFAULT_TYPE):
    """Ежедневная публикация сообщения верификации утром (удаляет предыдущее)."""
    try:
        await _post_verify(context.bot)
        runs = load_last_daily_run()
        runs["verify"] = datetime.date.today().isoformat()
        save_last_daily_run(runs)
        print("[daily_verify] Сообщение верификации обновлено")
    except Exception as e:
        print(f"[daily_verify] Ошибка: {e}")


async def _post_verify(bot, thread_id: int | None = None) -> int | None:
    """Удаляет предыдущее сообщение верификации и публикует новое. Возвращает ID нового сообщения."""
    if thread_id is None:
        thread_id = VERIFY_THREAD_ID or STATS_THREAD_ID

    me = await bot.get_me()
    bot_url = f"https://t.me/{me.username}?start=verify"

    text = (
        "👋 <b>Соседи, привет!</b>\n\n"
        "Если вы ещё не прошли верификацию — займёт меньше минуты.\n"
        "Укажите свой этаж и квартиру, чтобы соседи знали кто вы.\n\n"
        "Нажмите кнопку ниже 👇"
    )
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Пройти верификацию", url=bot_url)]
    ])

    kwargs = dict(chat_id=CHAT_ID, text=text, parse_mode="HTML", reply_markup=keyboard)
    if thread_id:
        kwargs["message_thread_id"] = thread_id

    # Если тема закрыта — сначала открываем (нужно и для удаления, и для отправки)
    reopened = False
    if thread_id:
        try:
            await bot.reopen_forum_topic(chat_id=CHAT_ID, message_thread_id=thread_id)
            reopened = True
        except Exception:
            pass

    # Удаляем предыдущее сообщение (тема уже открыта)
    prev_id = load_last_verify_msg()
    if prev_id:
        try:
            await bot.delete_message(chat_id=CHAT_ID, message_id=prev_id)
        except Exception as e:
            print(f"[_post_verify] Не удалось удалить пред. сообщение {prev_id}: {e}")

    msg = await bot.send_message(**kwargs)
    save_last_verify_msg(msg.message_id)

    # Удаляем сервисное сообщение "тема открыта" (оно идёт прямо перед нашим)
    if reopened:
        try:
            await bot.delete_message(chat_id=CHAT_ID, message_id=msg.message_id - 1)
        except Exception:
            pass

    if reopened:
        try:
            await bot.close_forum_topic(chat_id=CHAT_ID, message_thread_id=thread_id)
        except Exception:
            pass
        # Удаляем сервисное сообщение "тема закрыта" (оно идёт прямо после нашего)
        try:
            await bot.delete_message(chat_id=CHAT_ID, message_id=msg.message_id + 1)
        except Exception:
            pass

    return msg.message_id


async def postverify_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Публикует в чат приглашение пройти верификацию (удаляет предыдущее).

    /postverify        — в тему VERIFY_THREAD_ID (или STATS_THREAD_ID)
    /postverify 1234   — в конкретную тему с ID 1234
    """
    if update.effective_user.id not in ADMIN_IDS:
        return

    thread_id = None
    if context.args and context.args[0].isdigit():
        thread_id = int(context.args[0])

    try:
        msg_id = await _post_verify(context.bot, thread_id=thread_id)
        used_thread = thread_id or VERIFY_THREAD_ID or STATS_THREAD_ID
        tinfo = f" (тема {used_thread})" if used_thread else ""
        await update.message.reply_text(f"✅ Сообщение верификации опубликовано{tinfo}. ID: {msg_id}")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = load_db()
    if not db:
        await update.message.reply_text("База пустая.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Жильцы"

    # Заголовки
    headers = ["Этаж", "Квартира", "Имя", "Telegram", "Ссылка"]
    header_fill = PatternFill("solid", fgColor="2D6A9F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Сортировка: сначала по этажу, потом по квартире
    rows = sorted(db.values(), key=lambda x: (x.get("floor", 0), x.get("flat", 0)))

    fill_alt = PatternFill("solid", fgColor="EBF3FA")
    for i, r in enumerate(rows, 2):
        fill = fill_alt if i % 2 == 0 else None
        for col, val in enumerate([
            r.get("floor", ""),
            r.get("flat", ""),
            r.get("name", ""),
            r.get("telegram", "нет"),
            r.get("telegram_link", ""),
        ], 1):
            cell = ws.cell(row=i, column=col, value=val)
            if fill:
                cell.fill = fill

    # Ширина колонок
    for col, width in zip("ABCDE", [8, 10, 30, 22, 35]):
        ws.column_dimensions[col].width = width

    # Сохраняем в буфер
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.date.today().strftime("%d.%m.%Y")
    filename = f"жильцы_{today}.xlsx"

    await update.message.reply_document(
        document=buf,
        filename=filename,
        caption=f"📊 База жильцов на {today} — {len(rows)} записей"
    )


async def backup_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Создаёт ZIP-архив со всеми данными бота и отправляет администратору."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    import io, zipfile

    bot_dir = os.path.dirname(__file__)
    files_to_backup = [
        "residents.json",
        "unregistered.json",
        "join_requests.json",
        "digest_archive.json",
        "daily_messages.json",
        "last_stats_msg.json",
        "house.jpg",
        "nudge_photo.jpg",
    ]

    buf = io.BytesIO()
    today = datetime.date.today().strftime("%d.%m.%Y")
    added = []

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname in files_to_backup:
            path = os.path.join(bot_dir, fname)
            if os.path.exists(path):
                zf.write(path, fname)
                added.append(fname)

    buf.seek(0)
    await update.message.reply_document(
        document=buf,
        filename=f"bot_backup_{today}.zip",
        caption=f"🗂 Резервная копия на {today}\n📁 Файлов: {len(added)}\n\n" + "\n".join(f"• {f}" for f in added),
    )


async def unregistered_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return

    unreg = load_unregistered()
    if not unreg:
        await update.message.reply_text("✅ Все кто пишет в чате — есть в базе.")
        return

    lines = [f"👻 <b>Незарегистрированные участники</b> — {len(unreg)} чел.\n"]
    keyboard_rows = []
    for uid, u in sorted(unreg.items(), key=lambda x: -x[1].get("count", 0)):
        name = u.get("name", "Неизвестный")
        username = u.get("username") or ""
        user_id_val = u.get("user_id") or uid
        count = u.get("count", 0)
        first = u.get("first_seen", "")
        last = u.get("last_seen", "")
        # Кликабельная ссылка: по username если есть, иначе по ID
        if username:
            uname_clean = username.lstrip("@")
            link = f"https://t.me/{uname_clean}"
        else:
            link = f"tg://user?id={user_id_val}"
        name_link = f"<a href=\"{link}\">{name}</a>"
        tg = f" ({username})" if username else ""
        lines.append(f"• {name_link}{tg} — {count} сообщ., с {first} по {last}")
        keyboard_rows.append([InlineKeyboardButton(
            f"✓ убрать {name[:20]}",
            callback_data=f"unreg_dismiss:{uid}"
        )])

    keyboard_rows.append([InlineKeyboardButton("✓ убрать всех", callback_data="unreg_dismiss_all")])
    await update.message.reply_text(
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(keyboard_rows)
    )


NUDGE_PHOTO_FILE = os.path.join(os.path.dirname(__file__), "nudge_photo.jpg")


async def nudge_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправляет незарегистрированным участникам приглашение пройти верификацию.

    Использование:
      /nudge              — всем незарегистрированным, стандартный текст
      /nudge 348          — только жильцу кв.348 (из базы или незарег.)
      /nudge Свой текст   — всем, со своим текстом
      /nudge 348 Текст    — конкретной квартире со своим текстом
    """
    if update.effective_user.id not in ADMIN_IDS:
        return

    me = await context.bot.get_me()
    bot_url = f"https://t.me/{me.username}?start=verify"

    DEFAULT_TEXT = (
        "Привет! 👋\n\n"
        "Вы участник чата жильцов дома 6.2.2 "
        "(д.1Вк8 квартал №160 | 7 подъезд).\n\n"
        "Чтобы соседи знали кто вы — пройдите короткое подтверждение квартиры."
    )

    args = list(context.args) if context.args else []

    # Определяем: первый аргумент — квартира или текст?
    target_flat = None
    if args and args[0].isdigit():
        target_flat = int(args.pop(0))

    msg_text = " ".join(args).strip() if args else DEFAULT_TEXT

    # Кнопки верификации прямо в чате (только для стандартного текста)
    nudge_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Пройти верификацию", callback_data="yes_house")],
        [InlineKeyboardButton("❌", callback_data="no_house")],
    ]) if msg_text == DEFAULT_TEXT else None

    # Собираем список получателей
    recipients = {}  # uid -> name

    if target_flat:
        db = load_db()
        for rec in db.values():
            if rec.get("flat") == target_flat and rec.get("user_id"):
                uid = rec["user_id"]
                recipients[uid] = rec.get("name", str(uid))
        if not recipients:
            await update.message.reply_text(
                f"❌ Нет жильцов с user_id для кв. {target_flat}."
            )
            return
    else:
        unreg = load_unregistered()
        if not unreg:
            await update.message.reply_text("✅ Нет незарегистрированных — некому писать.")
            return
        for uid, u in unreg.items():
            recipients[int(uid)] = u.get("name", uid)

    # Фото для рассылки: сначала nudge_photo.jpg, иначе house.jpg
    photo_file = None
    if os.path.exists(NUDGE_PHOTO_FILE):
        photo_file = NUDGE_PHOTO_FILE
    elif os.path.exists(HOUSE_PHOTO):
        photo_file = HOUSE_PHOTO

    sent, failed = 0, 0
    for uid, name in recipients.items():
        try:
            if photo_file:
                with open(photo_file, "rb") as photo:
                    await context.bot.send_photo(
                        chat_id=uid,
                        photo=photo,
                        caption=msg_text,
                        reply_markup=nudge_keyboard,
                    )
            else:
                await context.bot.send_message(
                    chat_id=uid,
                    text=msg_text,
                    reply_markup=nudge_keyboard,
                )
            sent += 1
            await asyncio.sleep(0.3)
        except Exception:
            failed += 1

    target_str = f"кв. {target_flat}" if target_flat else "всем незарегистрированным"
    await update.message.reply_text(
        f"📨 Отправлено {target_str}: {sent} чел."
        + (f"\n❌ Не доставлено: {failed}" if failed else "")
    )


async def nudge_set_photo_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Установить фото для рассылки /nudge. Пришли фото следующим сообщением."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    waiting_nudge_photo.add(update.effective_user.id)
    await update.message.reply_text(
        "📷 Отправь фото следующим сообщением — оно будет прикрепляться к nudge-рассылке.\n"
        "Чтобы убрать фото — /nudgeclearphoto"
    )


async def nudge_clear_photo_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Убрать фото из nudge-рассылки."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    if os.path.exists(NUDGE_PHOTO_FILE):
        os.remove(NUDGE_PHOTO_FILE)
        await update.message.reply_text("✅ Фото для рассылки удалено.")
    else:
        await update.message.reply_text("Фото не было установлено.")


async def del_user_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return

    if not context.args:
        await update.message.reply_text(
            "Использование: <code>/deluser номер_квартиры</code>\n"
            "Пример: <code>/deluser 348</code>",
            parse_mode="HTML"
        )
        return

    try:
        flat = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Укажите номер квартиры числом.")
        return

    db = load_db()
    residents = [(k, v) for k, v in db.items() if v.get("flat") == flat]

    if not residents:
        await update.message.reply_text(f"❌ Квартира {flat} не найдена в базе.")
        return

    lines = [f"🚪 <b>Квартира {flat}</b> — найдено записей: {len(residents)}\n"]
    for _, v in residents:
        tg = v.get("telegram", "нет")
        lines.append(f"👤 {v.get('name', '—')}  |  {tg}")

    lines.append("\nУдалить?")

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("🗑 Удалить", callback_data=f"delconfirm_{flat}"),
        InlineKeyboardButton("❌ Отмена", callback_data=f"delcancel_{flat}"),
    ]])
    await update.message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=keyboard)


async def flat_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /flat <номер> — показывает все профили с этой квартирой и кнопки удаления."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    if not context.args:
        await update.message.reply_text(
            "Использование: <code>/flat 348</code>\n"
            "Показывает все профили с этой квартирой.",
            parse_mode="HTML"
        )
        return

    try:
        flat = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Укажите номер квартиры числом.")
        return

    db = load_db()
    residents = [(k, v) for k, v in db.items() if v.get("flat") == flat]

    if not residents:
        await update.message.reply_text(f"❌ Квартира {flat} — в базе не найдено.")
        return

    await update.message.reply_text(
        f"🚪 <b>Квартира {flat}</b> — найдено записей: {len(residents)}",
        parse_mode="HTML"
    )

    for key, v in residents:
        name = v.get("name", "—")
        floor = v.get("floor", "?")
        tg = v.get("telegram", "нет")
        tg_link = v.get("telegram_link", "")
        uid = v.get("user_id", "")
        name_linked = f'<a href="{tg_link}">{name}</a>' if tg_link else name

        text = (
            f"👤 {name_linked}\n"
            f"🏢 Этаж: {floor}  🚪 Квартира: {flat}\n"
            f"🔗 {tg}"
            + (f"\n🆔 user_id: {uid}" if uid else "")
        )
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("🗑 Удалить профиль", callback_data=f"flatdel_{key}"),
        ]])
        await update.message.reply_text(text, parse_mode="HTML", reply_markup=keyboard)


def parse_floor_flat(text: str):
    """Извлекает этаж и квартиру из произвольного текста."""
    import re
    text_lower = text.lower()

    # Ищем этаж: "3 этаж", "этаж 3", "3эт", "эт.3"
    floor = None
    m = re.search(r'(\d+)\s*эт(?:аж)?', text_lower)
    if m:
        floor = int(m.group(1))
    else:
        m = re.search(r'эт(?:аж)?[\s\.\-]*(\d+)', text_lower)
        if m:
            floor = int(m.group(1))

    # Ищем квартиру: "327 квартира", "квартира 327", "кв.327", "кв 327"
    flat = None
    m = re.search(r'(\d+)\s*кв(?:арт(?:ира)?)?(?:\b|\.)', text_lower)
    if m:
        flat = int(m.group(1))
    else:
        m = re.search(r'кв(?:арт(?:ира)?)?[\s\.\-]*(\d+)', text_lower)
        if m:
            flat = int(m.group(1))

    # Fallback: любое число в диапазоне 312–532 считаем квартирой
    if not flat:
        for num_str in re.findall(r'\b(\d{3})\b', text_lower):
            n = int(num_str)
            if 312 <= n <= 532:
                flat = n
                break

    # Если этаж не указан — определяем по квартире из floor_flats
    if flat and not floor:
        for fl, flats in floor_flats.items():
            if flat in flats:
                floor = fl
                break

    return floor, flat


async def handle_new_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Когда кто-то вступает в группу — добавляем в незарегистрированных если не в базе,
    или ставим тег если уже верифицирован."""
    msg = update.message
    if not msg or not msg.new_chat_members:
        return
    for user in msg.new_chat_members:
        if user.is_bot:
            continue

        uname = f"@{user.username}" if user.username else None
        user_link = f"https://t.me/{user.username}" if user.username else f"tg://user?id={user.id}"
        link_display = uname or f"tg://user?id={user.id}"

        if is_in_db(user.id):
            # Уже в базе — снимаем флаг left, ставим тег и уведомляем
            db = load_db()
            flat, floor = 0, 0
            for k, rec in db.items():
                if rec.get("user_id") == user.id:
                    flat = rec.get("flat", 0)
                    floor = rec.get("floor", 0)
                    if rec.get("left"):
                        db[k].pop("left", None)
                        save_db(db)
                    if flat:
                        tag = f"{floor} эт {flat}" if floor else str(flat)
                        _kbd = InlineKeyboardMarkup([[
                            InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{user.id}_{floor}_{flat}"),
                            InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{user.id}"),
                        ]])
                        try:
                            await context.bot.send_message(
                                chat_id=ADMIN_ID,
                                text=f"🔄 <b>Жилец вернулся в чат</b>\n👤 <a href=\"{user_link}\">{user.full_name}</a>, кв. {flat}\n\nПоставить подпись <b>«{tag}»</b>?",
                                parse_mode="HTML",
                                reply_markup=_kbd,
                            )
                        except Exception:
                            pass
                    break
            try:
                await context.bot.send_message(
                    chat_id=ADMIN_ID,
                    text=(
                        f"✅ <b>Новый жилец записан в базу</b>\n\n"
                        f"👤 <a href=\"{user_link}\">{user.full_name}</a>\n"
                        f"🏢 Этаж: {floor}\n"
                        f"🚪 Квартира: {flat}\n"
                        f"🔗 {link_display}"
                    ),
                    parse_mode="HTML",
                )
            except Exception:
                pass
            continue

        uid = str(user.id)
        unreg = load_unregistered()
        if uid not in unreg:
            unreg[uid] = {
                "name": user.full_name,
                "username": uname,
                "user_id": user.id,
                "first_seen": datetime.date.today().isoformat(),
                "last_seen": datetime.date.today().isoformat(),
                "count": 0,
            }
            save_unregistered(unreg)

        try:
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=(
                    f"⚠️ <b>Вступил в группу (не в базе)</b>\n\n"
                    f"👤 <a href=\"{user_link}\">{user.full_name}</a>\n"
                    f"🔗 {link_display}\n"
                    f"ID: <code>{user.id}</code>"
                ),
                parse_mode="HTML",
            )
        except Exception:
            pass


async def handle_left_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ChatMemberHandler — срабатывает когда участник покидает супергруппу."""
    cmu = update.chat_member
    if not cmu or cmu.chat.id != CHAT_ID:
        return

    old_status = cmu.old_chat_member.status
    new_status = cmu.new_chat_member.status

    # Интересуют только переходы из участника в «вышел» или «кикнут»
    was_member = old_status in ("member", "administrator", "restricted")
    is_gone = new_status in ("left", "kicked")
    if not (was_member and is_gone):
        return

    user = cmu.new_chat_member.user
    if user.is_bot:
        return

    uname = f"@{user.username}" if user.username else "нет username"
    user_link = f"https://t.me/{user.username}" if user.username else f"tg://user?id={user.id}"

    db = load_db()
    in_db = False
    for k, v in db.items():
        if v.get("user_id") == user.id:
            in_db = True
            db[k]["left"] = True
            save_db(db)
            break
    status_label = "📋 есть в базе жильцов" if in_db else "❓ не в базе"
    action = "🚫 Кикнут" if new_status == "kicked" else "🚪 Покинул чат"

    try:
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=(
                f"{action}\n\n"
                f"👤 <a href=\"{user_link}\">{user.full_name}</a>\n"
                f"Username: {uname}\n"
                f"ID: <code>{user.id}</code>\n"
                f"{status_label}"
            ),
            parse_mode="HTML",
        )
    except Exception:
        pass


async def handle_intro_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Когда кто-то пишет в тред Знакомства — парсим этаж/квартиру и обновляем базу."""
    if not INTRO_THREAD_ID:
        return
    msg = update.message
    if not msg or msg.chat.id != CHAT_ID:
        return
    if getattr(msg, "message_thread_id", None) != INTRO_THREAD_ID:
        return

    user = msg.from_user
    if not user or user.is_bot:
        return

    user_id = user.id
    username = f"@{user.username}" if user.username else None
    full_name = user.full_name
    tg_link = f"https://t.me/{user.username}" if user.username else f"tg://user?id={user_id}"
    tg_text = username or str(user_id)

    db = load_db()

    # Ищем существующую запись пользователя (по user_id или @username)
    existing_key = None
    if str(user_id) in db:
        existing_key = str(user_id)
    elif any(v.get("user_id") == user_id for v in db.values()):
        existing_key = next(k for k, v in db.items() if v.get("user_id") == user_id)
    elif username:
        for k, v in db.items():
            if v.get("telegram", "").lower() == username.lower():
                existing_key = k
                break

    # Если профиль уже есть в базе с квартирой — тихо обновляем данные и выходим
    if existing_key:
        existing_flat = db[existing_key].get("flat", 0)
        if existing_flat:
            db[existing_key]["user_id"] = user_id
            db[existing_key]["telegram"] = tg_text
            db[existing_key]["telegram_link"] = tg_link
            save_db(db)
            return

    # Пробуем распарсить этаж и квартиру из текста
    text = msg.text or msg.caption or ""
    floor, flat = parse_floor_flat(text)
    # Валидация: проверяем что этаж и квартира совпадают по схеме дома
    if floor and flat:
        if floor not in floor_flats or flat not in floor_flats.get(floor, []):
            floor, flat = None, None

    if existing_key:
        # В базе есть, но без квартиры — обновляем всё что знаем
        db[existing_key]["user_id"] = user_id
        db[existing_key]["telegram"] = tg_text
        db[existing_key]["telegram_link"] = tg_link
        if floor and flat:
            db[existing_key]["floor"] = floor
            db[existing_key]["flat"] = flat
        save_db(db)
        tag_flat = db[existing_key].get("flat", 0)
        tag_floor = db[existing_key].get("floor", 0)
    else:
        # Новый человек — добавляем с тем что знаем
        db[str(user_id)] = {
            "name": full_name,
            "floor": floor or 0,
            "flat": flat or 0,
            "telegram": tg_text,
            "telegram_link": tg_link,
            "user_id": user_id,
            "date_added": datetime.date.today().isoformat(),
        }
        save_db(db)
        tag_flat = flat or 0
        tag_floor = floor or 0

    # Спрашиваем администратора про тег если знаем квартиру
    if tag_flat:
        tag = f"{tag_floor} эт {tag_flat}" if tag_floor else str(tag_flat)
        name_in_db = db.get(existing_key or str(user_id), {}).get("name", full_name)
        tg_link_display = f"https://t.me/{user.username}" if user.username else f"tg://user?id={user_id}"
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{user_id}_{tag_floor}_{tag_flat}"),
            InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{user_id}"),
        ]])
        try:
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=(
                    f"📝 <b>Сообщение в треде знакомств</b>\n"
                    f"👤 <a href=\"{tg_link_display}\">{name_in_db}</a>, кв. {tag_flat}\n"
                    f"{username or str(user_id)}\n\n"
                    f"Поставить Telegram-подпись <b>«{tag}»</b>?"
                ),
                parse_mode="HTML",
                reply_markup=keyboard,
            )
        except Exception:
            pass


async def _enrich_user_id_from_message(user, context) -> bool:
    """Если пользователь есть в базе по @username без user_id — сохраняет ID и ставит тег.
    Если user_id уже есть, но telegram хранится как числовой — обновляет на @username.
    Возвращает True если запись была обновлена."""
    if not user or user.is_bot or not user.username:
        return False
    uname = f"@{user.username}".lower()
    uname_display = f"@{user.username}"
    tg_link_new = f"https://t.me/{user.username}"
    db_enriched = load_db()

    # Шаг 1: жилец уже с user_id, но telegram хранится как число — обновляем на @username
    for key, rec in db_enriched.items():
        if rec.get("user_id") != user.id:
            continue
        tg = str(rec.get("telegram", ""))
        if not tg.startswith("@"):
            db_enriched[key]["telegram"] = uname_display
            db_enriched[key]["telegram_link"] = tg_link_new
            save_db(db_enriched)
            flat = rec.get("flat", 0)
            floor = rec.get("floor", 0)
            try:
                await context.bot.send_message(
                    chat_id=ADMIN_ID,
                    text=(
                        f"🔗 <b>Обновлён профиль жильца</b>\n"
                        f"👤 <a href=\"{tg_link_new}\">{rec.get('name')}</a>, кв. {flat}\n"
                        f"Username: {uname_display}\n"
                        f"ID: <code>{user.id}</code>"
                    ),
                    parse_mode="HTML",
                )
            except Exception:
                pass
            return True

    for key, rec in db_enriched.items():
        if rec.get("user_id"):
            continue
        tg = str(rec.get("telegram", "")).lower()
        tg_link = str(rec.get("telegram_link", "")).lower()
        rec_uname = None
        if tg.startswith("@"):
            rec_uname = tg
        elif tg_link.startswith("https://t.me/"):
            rec_uname = "@" + tg_link.replace("https://t.me/", "").rstrip("/")
        if rec_uname and rec_uname == uname:
            db_enriched[key]["user_id"] = user.id
            save_db(db_enriched)
            flat = rec.get("flat", 0)
            floor = rec.get("floor", 0)
            tag = f"{floor} эт {flat}" if floor else f"{flat}"
            tg_link_display = f"https://t.me/{user.username}"
            keyboard = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton(f"🏷 Поставить тег «{tag}»", callback_data=f"settag_{user.id}_{floor}_{flat}"),
                    InlineKeyboardButton("❌ Не ставить", callback_data=f"skiptag_{user.id}"),
                ]
            ])
            try:
                await context.bot.send_message(
                    chat_id=ADMIN_ID,
                    text=(
                        f"✅ <b>Найден ID жильца</b>\n"
                        f"👤 <a href=\"{tg_link_display}\">{rec.get('name')}</a>, кв. {flat}\n"
                        f"Username: @{user.username}\n"
                        f"ID: <code>{user.id}</code>\n\n"
                        f"Поставить Telegram-подпись <b>«{tag}»</b>?"
                    ),
                    parse_mode="HTML",
                    reply_markup=keyboard,
                )
            except Exception:
                pass
            return True
    return False


async def handle_group_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Собирает сообщения группового чата для дневного дайджеста."""
    msg = update.message
    if not msg or not msg.text or msg.chat.id != CHAT_ID:
        return

    user = msg.from_user

    # Обогащение базы: если жилец с @username написал — запоминаем его числовой ID.
    # Работает для ЛЮБОГО треда, в том числе исключённых из дайджеста.
    await _enrich_user_id_from_message(user, context)

    # Исключаем треды статистики и вступления — они не нужны в дайджесте
    thread_id = getattr(msg, "message_thread_id", None)
    is_topic = getattr(msg, "is_topic_message", None)
    print(f"[msg_debug] id={msg.message_id} thread_id={thread_id} is_topic={is_topic} from={(user.full_name if user else '?')!r}")
    excluded_threads = {t for t in [STATS_THREAD_ID, INTRO_THREAD_ID, DIGEST_THREAD_ID, GENERAL_THREAD_ID] if t}
    if thread_id in excluded_threads:
        return

    sender = user.full_name if user else "Неизвестный"
    daily_messages.append({
        "name": sender,
        "text": msg.text[:300],
        "msg_id": msg.message_id,
        "thread_id": thread_id,
        "time": msg.date.strftime("%H:%M") if msg.date else "",
    })
    save_daily_messages(daily_messages)

    # Трекинг незарегистрированных: замечаем всех кто пишет, но не в базе
    if user and not user.is_bot and not is_in_db(user.id):
        uid = str(user.id)
        unreg = load_unregistered()
        if uid not in unreg:
            unreg[uid] = {
                "name": user.full_name,
                "username": f"@{user.username}" if user.username else None,
                "user_id": user.id,
                "first_seen": datetime.date.today().isoformat(),
                "last_seen": datetime.date.today().isoformat(),
                "count": 1,
            }
        else:
            unreg[uid]["last_seen"] = datetime.date.today().isoformat()
            unreg[uid]["count"] = unreg[uid].get("count", 0) + 1
            unreg[uid]["name"] = user.full_name
        save_unregistered(unreg)


async def build_digest_text(messages):
    """Строит текст дайджеста: AI выбирает главные темы за день со ссылками."""
    import re
    today = datetime.date.today().strftime("%-d %B %Y")
    count = len(messages)

    # Все сообщения с ID для AI
    prompt = "\n".join(
        f"[{m['time']}, ID:{m['msg_id']}] {m['name']}: {m['text']}"
        for m in messages
    )

    system = (
        "Ты помощник для чата жильцов многоквартирного дома. "
        "Твоя задача: найти ВСЕ отдельные ТЕМЫ разговора за день и перечислить их ВСЕ в хронологическом порядке. "
        "Тема — это группа сообщений об одном и том же вопросе. "
        "Несколько сообщений подряд об одном — это ОДНА тема, не несколько. "
        "НЕ пропускай темы, особенно те что появились позже — они тоже должны быть в списке. "
        "Если людей переписывались на 2 темы — выведи 2 пункта. "
        "Если на 5 тем — 5 пунктов. Не дроби одну тему на части.\n\n"
        "Для каждой темы:\n"
        "- Укажи ID ПЕРВОГО сообщения этой темы\n"
        "- Напиши ОДНО предложение с кратким содержанием всего диалога по этой теме. "
        "Упомяни итог или вывод если он был.\n\n"
        "Формат ответа (строго, без пояснений, без нумерации, каждая тема на отдельной строке):\n"
        "[ID:1234] Обсуждали проблему плесени и сырости — договорились подать коллективное обращение в УК\n"
        "[ID:1456] Михаил представил нового бота для чата жильцов\n"
        "[ID:1789] Соседи обсудили шум на лестнице по ночам"
    )

    try:
        client = OpenAI(api_key=OPENAI_API_KEY)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": f"Переписка за {today}:\n\n{prompt}"},
            ],
            max_tokens=1500,
        )
        ai_output = response.choices[0].message.content.strip()
    except Exception as e:
        return f"❌ Ошибка AI: {e}"

    lines_out = []
    for line in ai_output.splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.match(r'\[ID:(\d+)\]\s*(.*)', line)
        if match:
            msg_id = match.group(1)
            summary = match.group(2).strip()
            # Ищем thread_id для этого сообщения — нужен для мобильного Telegram
            thread_id_for_link = None
            for m in messages:
                if str(m.get("msg_id")) == msg_id:
                    thread_id_for_link = m.get("thread_id")
                    break
            # Трёхчастная ссылка для форум-групп: t.me/c/{chat}/{topic}/{msg}
            if thread_id_for_link and thread_id_for_link != 1:
                link = f"https://t.me/c/{CHAT_LINK_ID}/{thread_id_for_link}/{msg_id}"
            else:
                link = f"https://t.me/c/{CHAT_LINK_ID}/{msg_id}"
            lines_out.append(f'🔗 <a href="{link}">{summary}</a>')
        else:
            lines_out.append(f"• {line}")

    topics_text = "\n".join(lines_out) if lines_out else ai_output

    # Считаем новых жильцов зарегистрированных сегодня
    today_iso = datetime.date.today().isoformat()
    db_snap = load_db()
    new_neighbors = [
        v["name"] for v in db_snap.values()
        if v.get("date_added") == today_iso and v.get("flat")
    ]
    neighbors_line = ""
    if new_neighbors:
        neighbors_line = f"🏠 Новых соседей: {len(new_neighbors)}\n"

    return (
        f"📋 <b>Обзор чата за {today}</b>\n"
        f"💬 Сообщений за день: {count}\n"
        f"{neighbors_line}\n"
        f"{topics_text}"
    )


async def daily_digest(context: ContextTypes.DEFAULT_TYPE):
    """Запускается ежедневно в 23:50 — генерирует сводку и постит в чат."""
    global daily_messages

    if not daily_messages or not OPENAI_API_KEY:
        clear_daily_messages()
        daily_messages = []
        return

    text = await build_digest_text(daily_messages)

    # Сохраняем в архив перед очисткой
    today_iso = datetime.date.today().isoformat()
    save_digest_to_archive(today_iso, text)

    clear_daily_messages()
    daily_messages = []

    kwargs = dict(chat_id=CHAT_ID, text=text, parse_mode="HTML", disable_web_page_preview=True)
    if DIGEST_THREAD_ID:
        kwargs["message_thread_id"] = DIGEST_THREAD_ID
    try:
        await context.bot.send_message(**kwargs)
    except Exception:
        pass


async def digest_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /digest [дата] — сводка за сегодня или из архива за указанную дату.
    Работает из личного чата и из группового чата (только для админа)."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    # Если команда из группового чата — постим дайджест в тред дайджестов
    from_group = update.effective_chat.id == CHAT_ID

    async def _send(text, parse_mode="HTML"):
        """Всегда публикует дайджест в тред чата жильцов. Из группы — удаляет команду."""
        kwargs = dict(chat_id=CHAT_ID, text=text, parse_mode=parse_mode, disable_web_page_preview=True)
        if DIGEST_THREAD_ID:
            kwargs["message_thread_id"] = DIGEST_THREAD_ID
        await context.bot.send_message(**kwargs)
        if from_group:
            try:
                await update.message.delete()
            except Exception:
                pass
        else:
            await update.message.reply_text("✅ Дайджест опубликован в чате.", parse_mode=parse_mode)

    # Если передана дата — ищем в архиве
    if context.args:
        date_input = context.args[0].strip()
        parsed = None
        for fmt in ("%d.%m", "%d.%m.%Y", "%Y-%m-%d"):
            try:
                d = datetime.datetime.strptime(date_input, fmt)
                if fmt == "%d.%m":
                    d = d.replace(year=datetime.date.today().year)
                parsed = d.date()
                break
            except ValueError:
                continue

        if not parsed:
            await _send("❌ Не понял дату. Примеры: <code>/digest 20.05</code> или <code>/digest 20.05.2025</code>")
            return

        archive = load_digest_archive()
        key = parsed.isoformat()
        if key in archive:
            await _send(archive[key])
        else:
            dates = sorted(archive.keys(), reverse=True)
            if dates:
                dates_str = "\n".join(
                    f"• {datetime.date.fromisoformat(d).strftime('%d.%m.%Y')}"
                    for d in dates[:20]
                )
                await _send(
                    f"📭 Дайджест за <b>{parsed.strftime('%d.%m.%Y')}</b> не найден.\n\n"
                    f"📚 Доступные даты:\n{dates_str}"
                )
            else:
                await _send("📭 Архив дайджестов пуст.")
        return

    # Без аргумента — генерируем за сегодня
    if not daily_messages:
        archive = load_digest_archive()
        today_key = datetime.date.today().isoformat()
        if today_key in archive:
            await _send(archive[today_key])
        else:
            await _send("💬 Сообщений за сегодня пока нет.")
        return

    if not OPENAI_API_KEY:
        await _send("❌ OpenAI ключ не настроен.")
        return

    if not from_group:
        await update.message.reply_text("⏳ Генерирую сводку...")
    text = await build_digest_text(daily_messages)
    save_digest_to_archive(datetime.date.today().isoformat(), text)
    await _send(text)


async def digest_private_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /digestme — присылает дайджест только в личку администратору."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    if not daily_messages:
        archive = load_digest_archive()
        today_key = datetime.date.today().isoformat()
        if today_key in archive:
            await context.bot.send_message(chat_id=ADMIN_ID, text=archive[today_key], parse_mode="HTML")
        else:
            await context.bot.send_message(chat_id=ADMIN_ID, text="💬 Сообщений за сегодня пока нет.")
        return

    if not OPENAI_API_KEY:
        await context.bot.send_message(chat_id=ADMIN_ID, text="❌ OpenAI ключ не настроен.")
        return

    await context.bot.send_message(chat_id=ADMIN_ID, text="⏳ Генерирую сводку...")
    text = await build_digest_text(daily_messages)
    await context.bot.send_message(chat_id=ADMIN_ID, text=text, parse_mode="HTML")


async def set_photo_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        return
    waiting_photo.add(update.effective_user.id)
    await update.message.reply_text("📷 Отправьте фото дома следующим сообщением.")


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in ADMIN_IDS:
        return

    photo = update.message.photo[-1]
    file = await context.bot.get_file(photo.file_id)

    if user_id in waiting_nudge_photo:
        await file.download_to_drive(NUDGE_PHOTO_FILE)
        waiting_nudge_photo.discard(user_id)
        await update.message.reply_text("✅ Фото для рассылки сохранено. Теперь /nudge будет отправлять его.")
    elif user_id in waiting_photo:
        await file.download_to_drive(HOUSE_PHOTO)
        waiting_photo.discard(user_id)
        await update.message.reply_text("✅ Фото дома сохранено и будет показываться при проверке жильцов.")


async def join_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.chat_join_request.from_user
    chat_id = update.chat_join_request.chat.id
    join_requests[user.id] = chat_id
    save_join_requests(join_requests)
    # Сохраняем имя и username для команды /pending
    try:
        meta = json.load(open(JOIN_REQUESTS_META_FILE)) if os.path.exists(JOIN_REQUESTS_META_FILE) else {}
    except Exception:
        meta = {}
    meta[str(user.id)] = {
        "name": user.full_name,
        "username": f"@{user.username}" if user.username else None,
    }
    with open(JOIN_REQUESTS_META_FILE, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    try:
        await send_welcome(user.id, context)
    except Exception:
        pass


async def delete_forum_topic_service_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Автоматически удаляет сервисные сообщения об открытии/закрытии темы."""
    try:
        await update.message.delete()
    except Exception:
        pass


async def approve_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/approve <user_id> — вручную одобрить заявку на вступление."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    if not context.args or not context.args[0].isdigit():
        await update.message.reply_text("Использование: /approve <user_id>")
        return
    target_id = int(context.args[0])
    req_chat_id = join_requests.get(target_id, CHAT_ID)
    try:
        await context.bot.approve_chat_join_request(chat_id=req_chat_id, user_id=target_id)
        join_requests.pop(target_id, None)
        save_join_requests(join_requests)
        try:
            await context.bot.send_message(
                chat_id=target_id,
                text="✅ Заявка принята, добро пожаловать!"
            )
        except Exception:
            pass
        await update.message.reply_text(f"✅ Заявка пользователя {target_id} одобрена.")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


async def reject_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/reject <user_id> — вручную отклонить заявку на вступление."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    if not context.args or not context.args[0].isdigit():
        await update.message.reply_text("Использование: /reject <user_id>")
        return
    target_id = int(context.args[0])
    req_chat_id = join_requests.get(target_id, CHAT_ID)
    try:
        await context.bot.decline_chat_join_request(chat_id=req_chat_id, user_id=target_id)
        join_requests.pop(target_id, None)
        save_join_requests(join_requests)
        try:
            await context.bot.send_message(chat_id=target_id, text="❌ Ваша заявка отклонена.")
        except Exception:
            pass
        await update.message.reply_text(f"✅ Заявка пользователя {target_id} отклонена.")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


async def pending_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/pending — список всех заявок на вступление с именами и кнопками."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    if not join_requests:
        await update.message.reply_text("✅ Активных заявок нет.")
        return
    try:
        meta = json.load(open(JOIN_REQUESTS_META_FILE)) if os.path.exists(JOIN_REQUESTS_META_FILE) else {}
    except Exception:
        meta = {}
    db = load_db()
    for uid_int, chat_id in join_requests.items():
        uid = str(uid_int)
        m = meta.get(uid, {})
        name = m.get("name") or "Неизвестно"
        username = m.get("username") or "—"
        # Ищем в базе жильцов
        in_db = None
        for v in db.values():
            if str(v.get("user_id", "")) == uid:
                in_db = v
                break
        db_info = f"✅ В базе: кв. {in_db['flat']}" if in_db and in_db.get("flat") else "❌ Не в базе"
        # Кликабельная ссылка: t.me/username если есть, иначе tg://user?id=
        raw_username = m.get("username") or ""
        raw_username_clean = raw_username.lstrip("@")
        if raw_username_clean:
            profile_link = f"https://t.me/{raw_username_clean}"
        else:
            profile_link = f"tg://user?id={uid}"
        name_link = f'<a href="{profile_link}">{name}</a>'
        username_display = f"@{raw_username_clean}" if raw_username_clean else "—"
        text = (
            f"👤 {name_link}  {username_display}\n"
            f"🆔 <code>{uid}</code>\n"
            f"{db_info}"
        )
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Принять", callback_data=f"approve_{uid}"),
                InlineKeyboardButton("❌ Отклонить", callback_data=f"reject_{uid}"),
            ],
            [InlineKeyboardButton("🔁 Отправить форму", callback_data=f"resend_{uid}")],
        ])
        await update.message.reply_text(text, parse_mode="HTML", reply_markup=keyboard)


async def resend_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/resend <user_id> — переотправить форму верификации конкретному пользователю."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    if not context.args or not context.args[0].isdigit():
        await update.message.reply_text("Использование: <code>/resend user_id</code>", parse_mode="HTML")
        return
    target_id = int(context.args[0])
    try:
        await send_welcome(target_id, context)
        await update.message.reply_text(f"✅ Форма верификации отправлена пользователю {target_id}.")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


async def ping_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Тест: проверяет что бот получает команды и может отвечать."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    await update.message.reply_text(
        f"✅ Бот работает\n"
        f"ADMIN_ID: <code>{ADMIN_ID}</code>\n"
        f"CHAT_ID: <code>{CHAT_ID}</code>",
        parse_mode="HTML",
    )
    print(f"[ping] от {update.effective_user.id}")


async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ссылка на веб-дашборд мониторинга бота."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    domains = os.environ.get("REPLIT_DOMAINS", "")
    domain = domains.split(",")[0].strip() if domains else ""
    if domain:
        url = f"https://{domain}/api/status"
        text = (
            f"📊 <b>Дашборд мониторинга</b>\n\n"
            f'<a href="{url}">{url}</a>\n\n'
            f"Статус бота, база жильцов, расписание, все команды с описаниями."
        )
    else:
        text = (
            "📊 <b>Дашборд мониторинга</b>\n\n"
            "Откройте <code>/api/status</code> на домене вашего проекта.\n"
            "Например: <code>https://ваш-проект.replit.app/api/status</code>"
        )
    await update.message.reply_text(text, parse_mode="HTML", disable_web_page_preview=False)
    print(f"[status] от {update.effective_user.id}")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Справка по всем командам администратора."""
    if update.effective_user.id not in ADMIN_IDS:
        return
    text = (
        "📖 <b>Команды бота «Подъезд 7»</b>\n\n"

        "🗃 <b>База жильцов</b>\n"
        "/residents — список всех жильцов по этажам\n"
        "/adduser @user этаж кв Имя — добавить вручную\n"
        "/deluser кв — удалить квартиру из базы\n"
        "/export — база в Excel-файл\n"
        "/backup — ZIP-архив всех данных\n\n"

        "📊 <b>Статистика</b>\n"
        "/stats — статистика в личку\n"
        "/poststats — опубликовать статистику в чат\n\n"

        "✅ <b>Верификация</b>\n"
        "/postverify [thread] — приглашение пройти верификацию\n"
        "/nudge [кв] [текст] — рассылка незарег. участникам\n"
        "/nudgephoto — установить фото для рассылки\n"
        "/nudgeclearphoto — убрать фото из рассылки\n"
        "/unregistered — список незарегистрированных\n\n"

        "🏷 <b>Теги участников</b>\n"
        "/settitles — проставить теги всем жильцам\n"
        "/cleartitles — снять все теги\n"
        "/tag кв тег — добавить метку квартире\n"
        "/untag кв тег — убрать метку\n"
        "/tags — все метки в базе\n"
        "/debugtitle — диагностика тегов\n\n"

        "📋 <b>Дайджест</b>\n"
        "/digest [дата] — опубликовать дайджест в чат\n"
        "/digestme — дайджест в личку\n\n"

        "🔗 <b>Привязка ID</b>\n"
        "/setid @user ID — вручную привязать user_id\n"
        "↩ переслать сообщение — автопривязка по username\n\n"

        "⚙️ <b>Настройки</b>\n"
        "/setphoto — установить фото дома\n"
        "/ping — проверка работы бота\n"
        "/status — ссылка на веб-дашборд мониторинга\n"
        "/help — эта справка"
    )
    await update.message.reply_text(text, parse_mode="HTML")


async def setid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/setid @username ID — вручную привязать user_id к жильцу по @username."""
    if update.effective_user.id not in ADMIN_IDS:
        return

    args = context.args
    if len(args) != 2:
        await update.message.reply_text(
            "Использование: <code>/setid @username ID</code>\n"
            "Пример: <code>/setid @Squalla 123456789</code>",
            parse_mode="HTML",
        )
        return

    raw_uname = args[0].lstrip("@").lower()
    raw_id = args[1]
    if not raw_id.lstrip("-").isdigit():
        await update.message.reply_text("❌ ID должен быть числом.")
        return

    new_uid = int(raw_id)
    db = load_db()
    matched_key = None
    matched_rec = None

    for key, rec in db.items():
        tg = str(rec.get("telegram", "")).lower().lstrip("@")
        tg_link = str(rec.get("telegram_link", "")).lower()
        rec_uname = None
        if rec.get("telegram", "").startswith("@"):
            rec_uname = tg
        elif tg_link.startswith("https://t.me/"):
            rec_uname = tg_link.replace("https://t.me/", "").rstrip("/")
        if rec_uname and rec_uname == raw_uname:
            matched_key = key
            matched_rec = rec
            break

    if not matched_key:
        await update.message.reply_text(f"❌ @{raw_uname} не найден в базе.")
        return

    db[matched_key]["user_id"] = new_uid
    save_db(db)

    flat = matched_rec.get("flat", 0)
    floor = matched_rec.get("floor", 0)
    tag = f"{floor} эт {flat}" if floor else str(flat)
    tag_ok = False
    if flat:
        tag_ok, _ = await _set_member_title(context.bot, new_uid, tag)
    tag_info = f"🏷 Подпись «{tag}» поставлена" if tag_ok else "⚠️ Подпись не удалось поставить"

    await update.message.reply_text(
        f"✅ Привязано!\n"
        f"👤 {matched_rec.get('name')}, кв. {flat}\n"
        f"@{raw_uname} → ID <code>{new_uid}</code>\n"
        f"{tag_info}",
        parse_mode="HTML",
    )
    print(f"[setid] @{raw_uname} → {new_uid} (кв.{flat})")


app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(CommandHandler("approve", approve_command))
app.add_handler(CommandHandler("reject", reject_command))
app.add_handler(CommandHandler("pending", pending_command))
app.add_handler(CommandHandler("resend", resend_command))
app.add_handler(CommandHandler("ping", ping_command))
app.add_handler(CommandHandler("status", status_command))
app.add_handler(CommandHandler("help", help_command))
app.add_handler(CommandHandler("setid", setid_command))
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("residents", residents_list))
app.add_handler(CommandHandler("stats", stats_command))
app.add_handler(CommandHandler("poststats", poststats_command))
app.add_handler(CommandHandler("postverify", postverify_command))
app.add_handler(CommandHandler("adduser", add_user_command))
app.add_handler(CommandHandler("deluser", del_user_command))
app.add_handler(CommandHandler("flat", flat_command))
app.add_handler(CommandHandler("export", export_command))
app.add_handler(CommandHandler("unregistered", unregistered_command))
app.add_handler(CommandHandler("nudge", nudge_command))
app.add_handler(CommandHandler("nudgephoto", nudge_set_photo_command))
app.add_handler(CommandHandler("nudgeclearphoto", nudge_clear_photo_command))
app.add_handler(CommandHandler("tag", tag_command))
app.add_handler(CommandHandler("untag", untag_command))
app.add_handler(CommandHandler("tags", tags_command))
app.add_handler(CommandHandler("tagall", tagall_command))
app.add_handler(CommandHandler("untagall", untagall_command))
app.add_handler(CommandHandler("settitles", settitles_command))
app.add_handler(CommandHandler("cleartitles", cleartitles_command))
app.add_handler(CommandHandler("debugtitle", debugtitle_command))
app.add_handler(CommandHandler("digest", digest_now_command))
app.add_handler(CommandHandler("digestme", digest_private_command))
app.add_handler(CommandHandler("setphoto", set_photo_command))
app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
app.add_handler(MessageHandler(filters.FORWARDED & ~filters.COMMAND, handle_forward))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
# Сборщик сообщений группы для дайджеста — group=1 чтобы не конфликтовать с handle_message
app.add_handler(MessageHandler(filters.Chat(CHAT_ID) & filters.TEXT, handle_group_message), group=1)
app.add_handler(MessageHandler(filters.Chat(CHAT_ID) & filters.TEXT, handle_intro_message), group=2)
app.add_handler(MessageHandler(filters.Chat(CHAT_ID) & filters.StatusUpdate.NEW_CHAT_MEMBERS, handle_new_member), group=3)
app.add_handler(ChatMemberHandler(handle_left_member, chat_member_types=ChatMemberHandler.CHAT_MEMBER), group=3)
app.add_handler(MessageHandler(
    filters.Chat(CHAT_ID) & (filters.StatusUpdate.FORUM_TOPIC_CREATED | filters.StatusUpdate.FORUM_TOPIC_CLOSED | filters.StatusUpdate.FORUM_TOPIC_REOPENED | filters.StatusUpdate.FORUM_TOPIC_EDITED),
    delete_forum_topic_service_message
), group=4)
app.add_handler(CallbackQueryHandler(button))
app.add_handler(ChatJoinRequestHandler(join_request))

STATUS_FILE = os.path.join(os.path.dirname(__file__), "status.json")

async def write_heartbeat(context: ContextTypes.DEFAULT_TYPE):
    """Пишет bot/status.json каждые 30 секунд — для дашборда мониторинга."""
    try:
        db = load_db()
        total = len([v for v in db.values() if v.get("flat")])
        with_uid = len([v for v in db.values() if v.get("user_id") and v.get("flat")])
        without_uid = total - with_uid
        left = len([v for v in db.values() if v.get("left")])
        unreg = load_unregistered()
        jreqs = load_join_requests()

        try:
            bot_username = context.bot.username or (await context.bot.get_me()).username
        except Exception:
            bot_username = ""

        status = {
            "timestamp": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "bot_alive": True,
            "bot_username": bot_username,
            "db": {
                "total_residents": total,
                "with_user_id": with_uid,
                "without_user_id": without_uid,
                "left_members": left,
                "unregistered_seen": len(unreg),
            },
            "features": {
                "openai": bool(OPENAI_API_KEY),
                "stats_thread": bool(STATS_THREAD_ID),
                "digest_thread": bool(DIGEST_THREAD_ID),
                "intro_thread": bool(INTRO_THREAD_ID),
                "verify_thread": bool(VERIFY_THREAD_ID),
                "general_thread": bool(GENERAL_THREAD_ID),
            },
            "daily_messages": len(daily_messages),
            "pending_join_requests": len(jreqs),
            "schedule": {
                "digest_utc": "20:50",
                "digest_msk": "23:50",
                "stats_utc": "07:00",
                "stats_msk": "10:00",
            },
            "commands": [
                "/residents", "/stats", "/poststats", "/postverify",
                "/adduser", "/deluser", "/export", "/backup",
                "/unregistered", "/digest", "/digestme",
                "/tag", "/untag", "/tags", "/settitles", "/cleartitles",
                "/nudge", "/nudgephoto", "/nudgeclearphoto",
                "/setid", "/setphoto", "/ping",
            ],
        }
        with open(STATUS_FILE, "w", encoding="utf-8") as f:
            json.dump(status, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[heartbeat] error: {e}")


# Хартбит каждые 30 секунд
app.job_queue.run_repeating(write_heartbeat, interval=30, first=5, name="heartbeat")

# Ежедневный дайджест в 23:50 МСК (20:50 UTC)
import datetime as _dt
_UTC = _dt.timezone.utc
app.job_queue.run_daily(
    daily_digest,
    time=_dt.time(hour=20, minute=50, second=0, tzinfo=_UTC),
    name="daily_digest",
)

# Ежедневная статистика в 10:00 МСК (07:00 UTC)
app.job_queue.run_daily(
    daily_stats,
    time=_dt.time(hour=7, minute=0, second=0, tzinfo=_UTC),
    name="daily_stats",
)

# Ежедневное обновление сообщения верификации в 10:02 МСК (07:02 UTC)
app.job_queue.run_daily(
    daily_verify,
    time=_dt.time(hour=7, minute=2, second=0, tzinfo=_UTC),
    name="daily_verify",
)

ADMIN_COMMANDS = [
    ("residents",       "📋 Список жильцов с квартирами"),
    ("stats",           "📊 Статистика чата"),
    ("poststats",       "📤 Опубликовать статистику в чат"),
    ("postverify",      "📌 Опубликовать форму верификации"),
    ("pending",         "⏳ Заявки на вступление"),
    ("approve",         "✅ Одобрить заявку — /approve <id>"),
    ("reject",          "❌ Отклонить заявку — /reject <id>"),
    ("adduser",         "➕ Добавить жильца — /adduser <кв> <имя>"),
    ("deluser",         "🗑 Удалить жильца — /deluser <кв>"),
    ("setid",           "🔗 Привязать Telegram ID к квартире"),
    ("flat",            "🏠 Инфо о квартире — /flat <номер>"),
    ("unregistered",    "👻 Незарегистрированные участники чата"),
    ("export",          "💾 Экспорт базы жильцов (JSON)"),
    ("digest",          "📋 Дайджест — /digest или /digest 01.06"),
    ("digestme",        "📬 Дайджест в личку"),
    ("tag",             "🏷 Поставить тег — /tag <кв> <тег>"),
    ("untag",           "🏷 Убрать тег — /untag <кв> <тег>"),
    ("tags",            "🏷 Список всех тегов"),
    ("settitles",       "✏️ Обновить подписи жильцов в чате"),
    ("cleartitles",     "🧹 Сбросить все подписи"),
    ("nudge",           "📣 Напомнить незарег. о верификации"),
    ("nudgephoto",      "🖼 Установить фото для nudge-сообщения"),
    ("nudgeclearphoto", "🗑 Убрать фото из nudge-сообщения"),
    ("setphoto",        "🖼 Установить фото дома для верификации"),
    ("ping",            "🏓 Проверить что бот живой"),
]


async def on_startup(application):
    """При старте — отправить форму верификации тем у кого висит заявка, но нет в базе.
    Также запускает пропущенные ежедневные задачи если бот не работал утром."""
    bot = application.bot

    # ── 0. Устанавливаем меню команд только для администраторов ──────────
    from telegram import BotCommand
    from telegram.constants import BotCommandScopeType
    try:
        bot_commands = [BotCommand(cmd, desc) for cmd, desc in ADMIN_COMMANDS]
        for admin_id in ADMIN_IDS:
            try:
                await bot.set_my_commands(
                    commands=bot_commands,
                    scope={"type": "chat", "chat_id": admin_id},
                )
            except Exception as e:
                print(f"[startup] set_my_commands for {admin_id}: {e}")
        # Обычным пользователям — пустое меню (не видят команды)
        await bot.set_my_commands(commands=[])
        print(f"[startup] Меню команд настроено для {len(ADMIN_IDS)} администраторов")
    except Exception as e:
        print(f"[startup] Ошибка настройки меню команд: {e}")

    # ── 1. Дозапуск пропущенных утренних задач ──────────────────────────
    today = datetime.date.today().isoformat()
    now_utc = datetime.datetime.now(datetime.timezone.utc)
    runs = load_last_daily_run()

    # Задачи запускаются в 07:00/07:02 UTC; если бот стартовал позже — догоняем
    if now_utc.hour >= 7:
        catchup_done = []
        if runs.get("stats") != today:
            try:
                await _post_stats(bot)
                runs["stats"] = today
                catchup_done.append("статистика")
            except Exception as e:
                print(f"[startup] catchup stats error: {e}")
        if runs.get("verify") != today:
            try:
                await _post_verify(bot)
                runs["verify"] = today
                catchup_done.append("верификация")
            except Exception as e:
                print(f"[startup] catchup verify error: {e}")
        if catchup_done:
            save_last_daily_run(runs)
            print(f"[startup] Дозапущено: {', '.join(catchup_done)}")
            try:
                await bot.send_message(
                    chat_id=ADMIN_ID,
                    text=f"⚡️ <b>Бот перезапущен</b> — дозапущены пропущенные задачи:\n" +
                         "\n".join(f"✅ {t}" for t in catchup_done),
                    parse_mode="HTML",
                )
            except Exception:
                pass

    # ── 2. Форма верификации тем у кого висит заявка ────────────────────
    db = load_db()
    registered_ids = {v.get("user_id") for v in db.values() if v.get("user_id")}
    pending = load_join_requests()
    sent = 0
    for uid_str, chat_id in pending.items():
        uid = int(uid_str)
        if uid not in registered_ids:
            try:
                if os.path.exists(HOUSE_PHOTO):
                    with open(HOUSE_PHOTO, "rb") as photo:
                        await bot.send_photo(
                            chat_id=uid,
                            photo=photo,
                            caption=WELCOME_TEXT,
                            parse_mode="HTML",
                            reply_markup=WELCOME_KEYBOARD,
                        )
                else:
                    await bot.send_message(
                        chat_id=uid,
                        text=WELCOME_TEXT,
                        parse_mode="HTML",
                        reply_markup=WELCOME_KEYBOARD,
                    )
                sent += 1
            except Exception:
                pass
    if sent:
        print(f"[startup] Отправлена форма {sent} пользователям с незавершённой заявкой")


app.post_init = on_startup

print("Бот запущен...")
app.run_polling(allowed_updates=Update.ALL_TYPES)
